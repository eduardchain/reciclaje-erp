"""
Migracion generica de organizaciones a EcoBalance.

Crea una nueva organizacion y carga snapshot al go-live a partir de un Excel:
- Maestros: Unidades de Negocio, Bodegas, categorias material/gasto/tercero,
  materiales, precios, terceros, cuentas.
- Saldos: initial_balance en cuentas y terceros (current_balance = initial_balance al crear).
- Stock: via POST /inventory/adjustments/increase/ (genera MaterialCostHistory).
- Activos fijos: con historical_load para los pre-sistema (sin proveedor ni cuenta)
  que no deben generar MoneyMovement ni mover balances.

Diferencias vs migrate_metarecycling.py (script anterior, hardcoded MetaRecycling):
- Unidades de Negocio se leen del Excel (hoja UnidadesNegocio) en vez de constante.
- Bodegas se leen del Excel (hoja Bodegas); fallback "Principal" si la hoja esta vacia.
- INVENTORY_REASON se construye dinamico: "Carga inicial migracion {org_name}".
- Default ThirdParty categories se obtienen de la API (GET /api/v1/third-party-categories/)
  despues de crear la org. NO se duplican.
- Nueva hoja "Precios" carga el catalogo de precios (POSTs separados compra/venta).
- CategoriaGastos soporta columnas asignacion_un, unidad_negocio, unidades_negocio (CSV).
- ActivosFijos soporta unidades_negocio (CSV) → applicable_business_unit_ids.
  Validacion client-side: depreciacion_acumulada <= valor_compra - valor_residual.
  Removida columna legacy "tasa_depreciacion" (se usa vida_util_meses exclusivamente).

Uso:
    cd backend
    ./venv/bin/python scripts/migrate_org.py \\
        --file ../data/migration_<org>.xlsx \\
        --org-name "Mi Empresa SAS" \\
        --superuser-email superadmin@ecobalance.com \\
        --superuser-password '<pwd>' \\
        --admin-email admin@miempresa.com \\
        --admin-name "Admin Owner" \\
        --dry-run

    # Aplicar:
    ./venv/bin/python scripts/migrate_org.py [...] --apply

Hojas del Excel (12; MapeoUN es read-only, las otras 11 escriben):
  1.  UnidadesNegocio       nombre, descripcion, is_active
  2.  Bodegas               nombre, descripcion, address, is_active
  3.  CategoriaMateriales   nombre, descripcion
  4.  CategoriaGastos       nombre, padre, es_directo, asignacion_un, unidad_negocio,
                            unidades_negocio (CSV), descripcion
  5.  CategoriaTerceros     nombre, tipo_comportamiento (solo raices), padre, descripcion
                            (extras; las 7 default vienen con la org)
  6.  MapeoUN               categoria_material, unidad_negocio_default, notas (read-only)
  7.  Materiales            codigo, nombre, categoria, unidad_negocio (override opcional),
                            unidad, descripcion
  8.  Precios               material_codigo, precio_compra, precio_venta, notas
                            (2 POSTs por fila a /price-lists/, uno por tipo)
  9.  Terceros              nombre, identificacion, email, telefono, direccion,
                            categorias (CSV), saldo_inicial, provision_type
 10.  Cuentas               nombre, tipo (cash/bank/digital), saldo_inicial,
                            numero_cuenta, banco
 11.  Inventario            material_codigo, bodega, cantidad, costo_unitario, fecha
 12.  ActivosFijos          nombre, codigo_activo, fecha_compra, fecha_inicio_depreciacion,
                            valor_compra, valor_residual, vida_util_meses,
                            depreciacion_acumulada, categoria_gasto, proveedor (XOR cuenta_pago),
                            cuenta_pago, unidad_negocio, unidades_negocio (CSV), notas
"""
import argparse
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Optional

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no esta instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

DEFAULT_WAREHOUSE_NAME = "Principal"  # fallback solo si hoja Bodegas esta vacia

ACCOUNT_TYPE_MAP = {
    "efectivo": "cash", "cash": "cash",
    "banco": "bank", "bank": "bank",
    "digital": "digital",
}

BEHAVIOR_TYPE_MAP = {
    "proveedor_material": "material_supplier",
    "proveedor_servicios": "service_provider",
    "cliente": "customer",
    "inversionista": "investor",
    "generico": "generic",
    "provision": "provision",
    "pasivo": "liability",
    "material_supplier": "material_supplier",
    "service_provider": "service_provider",
    "customer": "customer",
    "investor": "investor",
    "generic": "generic",
    "liability": "liability",
    "provision_": "provision",
}

ASIGNACION_UN_MAP = {
    "directo": "directo",
    "compartido": "compartido",
    "general": "general",
    "direct": "directo",
    "shared": "compartido",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_bool(value, default: bool = False) -> bool:
    if value is None or value == "":
        return default
    return str(value).strip().upper() in ("SI", "SÍ", "YES", "1", "TRUE", "X", "Y")


def parse_decimal(value, default="0") -> Decimal:
    if value is None or value == "":
        return Decimal(str(default))
    s = str(value).strip()
    if s in ("", "-", "$", "$ -"):
        return Decimal(str(default))
    try:
        return Decimal(s.replace("$", "").replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(str(default))


def parse_int(value, default=0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(parse_decimal(value, default))
    except Exception:
        return default


def parse_date(value) -> Optional[str]:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    s = str(value).strip()
    if not s:
        return None
    return s.split(" ")[0][:10]


def clean_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_csv(value) -> list[str]:
    """Parsea CSV simple (comma-separated), devuelve lista trimeada sin vacios."""
    s = clean_str(value)
    if not s:
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def read_sheet(wb, sheet_name: str) -> list[dict]:
    """Lee una hoja. Detecta header row automaticamente:
    - Si la fila 1 tiene una sola celda con texto que empieza con 'NOTAS' (template),
      header esta en fila 2 y datos desde fila 3.
    - Si no, header esta en fila 1.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    first = rows[0]
    is_notes_row = bool(
        first and first[0]
        and isinstance(first[0], str)
        and first[0].strip().upper().startswith("NOTAS")
        and all(v is None for v in first[1:])
    )
    header_idx = 1 if is_notes_row else 0
    if len(rows) < header_idx + 2:
        return []
    headers = [clean_str(h) or f"col_{i}" for i, h in enumerate(rows[header_idx])]
    headers = [h.lower().strip().rstrip("*").strip() for h in headers]
    result = []
    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if all(v is None for v in row):
            continue
        data = {"_row": row_idx}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                data[headers[col_idx]] = val
        result.append(data)
    return result


def find_in_map_ci(mapping: dict, key: Optional[str]) -> Optional[str]:
    if not key:
        return None
    key_lower = str(key).lower().strip()
    for k, v in mapping.items():
        if k.lower().strip() == key_lower:
            return v
    return None


def is_duplicate(status_code, resp) -> bool:
    if status_code == 409:
        return True
    resp_str = str(resp).lower()
    return status_code == 400 and (
        "ya existe" in resp_str or "duplicate" in resp_str or "already" in resp_str
    )


# ---------------------------------------------------------------------------
# API Client
# ---------------------------------------------------------------------------

class APIClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.token: Optional[str] = None
        self.org_id: Optional[str] = None
        self.session = requests.Session()

    def login(self, email: str, password: str) -> bool:
        resp = self.session.post(
            f"{self.base_url}/api/v1/auth/login/json",
            json={"email": email, "password": password},
        )
        if resp.status_code == 200:
            self.token = resp.json()["access_token"]
            self.session.headers.update({"Authorization": f"Bearer {self.token}"})
            return True
        print(f"  ERROR login: {resp.status_code} - {resp.text}")
        return False

    def set_org(self, org_id: str):
        self.org_id = org_id
        self.session.headers.update({"X-Organization-ID": org_id})

    def post(self, path: str, data: Optional[dict] = None):
        resp = self.session.post(f"{self.base_url}{path}", json=data or {})
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, resp.text

    def get(self, path: str, params: Optional[dict] = None):
        resp = self.session.get(f"{self.base_url}{path}", params=params or {})
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, resp.text

    def delete(self, path: str):
        resp = self.session.delete(f"{self.base_url}{path}")
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, resp.text


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

class Stats:
    def __init__(self, name: str):
        self.name = name
        self.created = 0
        self.skipped = 0
        self.errors = 0
        self.error_details: list[str] = []

    def add_error(self, msg: str):
        self.errors += 1
        self.error_details.append(msg)

    def report(self) -> str:
        parts = []
        if self.created:
            parts.append(f"{self.created} creados")
        if self.skipped:
            parts.append(f"{self.skipped} omitidos")
        if self.errors:
            parts.append(f"{self.errors} errores")
        return f"  [{self.name}] {', '.join(parts) if parts else 'vacio'}"


# ---------------------------------------------------------------------------
# Phase: org + admin
# ---------------------------------------------------------------------------

def reset_org(api: APIClient, org_name: str, dry_run: bool) -> bool:
    st, resp = api.get("/api/v1/system/organizations", {"include_inactive": "true"})
    if st != 200:
        print(f"  ERROR listando orgs: {st} - {resp}")
        return False
    existing = next(
        (o for o in resp if o.get("name", "").lower() == org_name.lower() and o.get("is_active")),
        None,
    )
    if not existing:
        print(f"  (--reset-org) Org '{org_name}' no existe o ya esta inactiva, skip")
        return True
    if dry_run:
        print(f"  (dry-run, --reset-org) Soft-deletaria org '{org_name}' ({existing['id']})")
        return True
    st, resp = api.delete(f"/api/v1/system/organizations/{existing['id']}")
    if st == 200:
        print(f"  (--reset-org) Org '{org_name}' desactivada ({existing['id']})")
        return True
    print(f"  ERROR soft-deleting org: {st} - {resp}")
    return False


def ensure_org_and_admin(
    api: APIClient, org_name: str, admin_email: str, admin_name: str, dry_run: bool
) -> Optional[str]:
    st, resp = api.get("/api/v1/system/organizations", {"include_inactive": "true"})
    if st != 200:
        print(f"  ERROR listando orgs: {st} - {resp}")
        return None
    # Ignorar orgs inactivas (residuos de --reset-org previos); el slug se autogenera
    # con sufijo cuando hace falta, asi que podemos crear una nueva con el mismo name.
    existing = next(
        (o for o in resp if o.get("name", "").lower() == org_name.lower() and o.get("is_active")),
        None,
    )
    if existing:
        print(
            f"  ORG '{org_name}' ya existe (id={existing['id']}, "
            f"miembros={existing.get('member_count', 0)}). Reusando."
        )
        return existing["id"]
    if dry_run:
        print(f"  (dry-run) Crearia org '{org_name}' + admin {admin_email}")
        return "00000000-0000-0000-0000-000000000000"
    payload = {"name": org_name, "admin_email": admin_email, "admin_full_name": admin_name}
    st, resp = api.post("/api/v1/system/organizations", payload)
    if st == 201:
        print(
            f"  ORG '{org_name}' creada (id={resp['id']}). "
            f"Admin: {admin_email} (password default '123456' si era usuario nuevo)."
        )
        return resp["id"]
    print(f"  ERROR creando org: {st} - {resp}")
    return None


# ---------------------------------------------------------------------------
# Phase: maestros simples
# ---------------------------------------------------------------------------

def fill_existing(api: APIClient, path: str, mapping: dict, key_field: str = "name"):
    st, resp = api.get(path, {"limit": 500})
    if st != 200:
        return
    items = resp.get("items", resp) if isinstance(resp, dict) else resp
    if isinstance(items, dict):
        items = items.get("items", [])
    for item in items or []:
        k = item.get(key_field)
        if k and k not in mapping:
            mapping[k] = item["id"]


def load_business_units(
    api: APIClient, rows: list[dict], dry_run: bool
) -> tuple[Stats, dict]:
    """Lee UnidadesNegocio del Excel y las crea via /api/v1/business-units/."""
    stats = Stats("UnidadesNegocio")
    mapping: dict[str, str] = {}
    if not rows:
        stats.add_error(
            "  Hoja UnidadesNegocio vacia o ausente. Las UNs son obligatorias para "
            "categorizar materiales/gastos/activos. Llena la hoja UnidadesNegocio del template."
        )
        return stats, mapping
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-bu-{nombre}"
            continue
        payload: dict = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        # is_active default true; solo enviar si esta explicitamente false
        is_active_raw = row.get("is_active")
        if is_active_raw is not None and is_active_raw != "":
            payload["is_active"] = parse_bool(is_active_raw, default=True)
        st, resp = api.post("/api/v1/business-units/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        fill_existing(api, "/api/v1/business-units/", mapping)
    return stats, mapping


def load_warehouses(
    api: APIClient, rows: list[dict], dry_run: bool
) -> tuple[Stats, dict]:
    """Lee Bodegas del Excel y las crea via /api/v1/warehouses/.
    Fallback: si la hoja esta vacia o ausente, crea una bodega 'Principal'.
    """
    stats = Stats("Bodegas")
    mapping: dict[str, str] = {}
    if not rows:
        # Fallback backwards-compat: una bodega 'Principal'
        nombre = DEFAULT_WAREHOUSE_NAME
        print(f"  (fallback) Hoja Bodegas vacia: creando bodega unica '{nombre}'")
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-wh-{nombre}"
            return stats, mapping
        st, resp = api.post("/api/v1/warehouses/", {"name": nombre})
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
            fill_existing(api, "/api/v1/warehouses/", mapping)
        else:
            stats.add_error(f"  Warehouse '{nombre}': {st} - {resp}")
        return stats, mapping

    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-wh-{nombre}"
            continue
        payload: dict = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        addr = clean_str(row.get("address")) or clean_str(row.get("direccion"))
        if addr:
            payload["address"] = addr
        is_active_raw = row.get("is_active")
        if is_active_raw is not None and is_active_raw != "":
            payload["is_active"] = parse_bool(is_active_raw, default=True)
        st, resp = api.post("/api/v1/warehouses/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        fill_existing(api, "/api/v1/warehouses/", mapping)
    return stats, mapping


def load_material_categories(api: APIClient, rows: list[dict], dry_run: bool) -> tuple[Stats, dict]:
    stats = Stats("CategoriaMateriales")
    mapping: dict[str, str] = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-mc-{nombre}"
            continue
        payload = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        st, resp = api.post("/api/v1/material-categories/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        fill_existing(api, "/api/v1/material-categories/", mapping)
    return stats, mapping


def load_expense_categories(
    api: APIClient,
    rows: list[dict],
    bu_map: dict,
    dry_run: bool,
) -> tuple[Stats, dict]:
    """Carga categorias de gasto en 2 pases (roots primero, despues hijos).
    Columnas extendidas:
      - asignacion_un ∈ {directo, compartido, general} (default general)
      - unidad_negocio (single) → default_business_unit_id
      - unidades_negocio (CSV) → default_applicable_business_unit_ids
    Si el padre tiene es_directo, la hija lo hereda (en row hija dejar es_directo vacio).
    """
    stats = Stats("CategoriaGastos")
    mapping: dict[str, str] = {}
    roots = [r for r in rows if not clean_str(r.get("padre"))]
    children = [r for r in rows if clean_str(r.get("padre"))]
    for row in roots + children:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        padre = clean_str(row.get("padre"))
        parent_id = find_in_map_ci(mapping, padre) if padre else None
        if padre and not parent_id and not dry_run:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): padre '{padre}' no encontrado")
            continue

        # Resolver asignacion_un
        asignacion_raw = clean_str(row.get("asignacion_un"))
        asignacion = None
        if asignacion_raw:
            asignacion = ASIGNACION_UN_MAP.get(asignacion_raw.lower().strip())
            if not asignacion:
                stats.add_error(
                    f"  Fila {row['_row']} ({nombre}): asignacion_un '{asignacion_raw}' "
                    f"invalida (esperado: directo|compartido|general)"
                )
                continue

        # Resolver default_business_unit_id (single)
        un_single = clean_str(row.get("unidad_negocio"))
        un_single_id = find_in_map_ci(bu_map, un_single) if un_single else None
        if un_single and not un_single_id and not dry_run:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): unidad_negocio '{un_single}' no encontrada"
            )
            continue

        # Resolver default_applicable_business_unit_ids (CSV)
        un_csv = parse_csv(row.get("unidades_negocio"))
        un_csv_ids: list[str] = []
        un_errors: list[str] = []
        for un_name in un_csv:
            un_id = find_in_map_ci(bu_map, un_name)
            if un_id:
                un_csv_ids.append(un_id)
            else:
                un_errors.append(un_name)
        if un_errors and not dry_run:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): unidades_negocio no encontradas: "
                f"{', '.join(un_errors)}"
            )
            continue

        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-ec-{nombre}"
            continue

        payload: dict = {
            "name": nombre,
            # is_direct_expense: roots usan valor del Excel; hijos heredan (campo dejado vacio)
            "is_direct_expense": parse_bool(row.get("es_directo")) if not padre else False,
        }
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        if parent_id:
            payload["parent_id"] = parent_id
        if asignacion:
            payload["default_allocation_type"] = asignacion
        if un_single_id:
            payload["default_business_unit_id"] = un_single_id
        if un_csv_ids:
            payload["default_applicable_business_unit_ids"] = un_csv_ids
        st, resp = api.post("/api/v1/expense-categories/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        st, resp = api.get("/api/v1/expense-categories/", {"limit": 500})
        if st == 200:
            for item in resp.get("items", []):
                if item["name"] not in mapping:
                    mapping[item["name"]] = item["id"]
    return stats, mapping


def fetch_default_tp_categories(api: APIClient, dry_run: bool) -> dict[str, str]:
    """Despues de crear la org, GET /api/v1/third-party-categories/ trae las 7
    categorias default seedadas automaticamente. Las usamos para construir el
    mapping y NUNCA duplicarlas en la fase de extras (hoja CategoriaTerceros).
    """
    mapping: dict[str, str] = {}
    if dry_run:
        # En dry-run inventamos placeholders para que el resto de la validacion siga
        for name in (
            "Proveedor Material",
            "Proveedor Servicios",
            "Cliente",
            "Socios",
            "Genérico",
            "Provisión",
            "Pasivo",
        ):
            mapping[name] = f"dry-run-tpc-default-{name}"
        return mapping
    fill_existing(api, "/api/v1/third-party-categories/", mapping)
    return mapping


def load_third_party_categories_extras(
    api: APIClient, rows: list[dict], existing_map: dict, dry_run: bool
) -> tuple[Stats, dict]:
    """Carga SOLO categorias extras del Excel (no las 7 default).
    Si una fila del Excel coincide con una default, se omite silenciosamente.
    2-pass: roots primero, despues hijos.
    """
    stats = Stats("CategoriaTerceros")
    # Empezamos con las defaults ya cargadas (no se cuentan como created)
    mapping: dict[str, str] = dict(existing_map)

    roots = [r for r in rows if not clean_str(r.get("padre"))]
    children = [r for r in rows if clean_str(r.get("padre"))]
    for row in roots + children:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        if find_in_map_ci(mapping, nombre):
            # Ya existe (default o creada previamente): no duplicar
            stats.skipped += 1
            continue
        padre = clean_str(row.get("padre"))
        parent_id = find_in_map_ci(mapping, padre) if padre else None
        bt_raw = clean_str(row.get("tipo_comportamiento"))
        behavior_type = None
        if bt_raw:
            behavior_type = BEHAVIOR_TYPE_MAP.get(bt_raw.lower().strip())
            if not behavior_type:
                stats.add_error(
                    f"  Fila {row['_row']} ({nombre}): tipo_comportamiento '{bt_raw}' invalido"
                )
                continue
        elif not padre:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): tipo_comportamiento obligatorio para raices"
            )
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-tpc-{nombre}"
            continue
        payload: dict = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        if behavior_type:
            payload["behavior_type"] = behavior_type
        if parent_id:
            payload["parent_id"] = parent_id
        st, resp = api.post("/api/v1/third-party-categories/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    return stats, mapping


# ---------------------------------------------------------------------------
# Phase: MapeoUN + Materiales
# ---------------------------------------------------------------------------

def build_un_map(rows: list[dict]) -> dict[str, str]:
    """Construye categoria_material → unidad_negocio_default."""
    result = {}
    for row in rows:
        cat = clean_str(row.get("categoria_material"))
        un = clean_str(row.get("unidad_negocio_default"))
        if cat and un:
            result[cat] = un
    return result


def load_materials(
    api: APIClient,
    rows: list[dict],
    cat_map: dict,
    bu_map: dict,
    un_by_category: dict[str, str],
    dry_run: bool,
) -> tuple[Stats, dict]:
    stats = Stats("Materiales")
    mapping: dict[str, str] = {}
    for row in rows:
        codigo = clean_str(row.get("codigo"))
        nombre = clean_str(row.get("nombre"))
        categoria = clean_str(row.get("categoria"))
        un_override = clean_str(row.get("unidad_negocio"))
        errors = []
        if not codigo:
            errors.append("codigo obligatorio")
        if not nombre:
            errors.append("nombre obligatorio")
        cat_id = find_in_map_ci(cat_map, categoria) if categoria else None
        if not categoria:
            errors.append("categoria obligatoria")
        elif not cat_id:
            errors.append(f"categoria '{categoria}' no encontrada")
        # Resolver UN: override de la fila, o default del MapeoUN por categoria
        un_name = un_override or (un_by_category.get(categoria) if categoria else None)
        un_id = find_in_map_ci(bu_map, un_name) if un_name else None
        if not un_name:
            errors.append(
                f"unidad_negocio no resuelta (sin override y categoria '{categoria}' no esta en MapeoUN)"
            )
        elif not un_id:
            errors.append(f"unidad_negocio '{un_name}' no encontrada")
        if errors:
            stats.add_error(f"  Fila {row['_row']} ({codigo or '?'}): {'; '.join(errors)}")
            continue
        if dry_run:
            stats.created += 1
            mapping[codigo] = f"dry-run-mat-{codigo}"
            continue
        payload = {
            "code": codigo,
            "name": nombre,
            "category_id": cat_id,
            "business_unit_id": un_id,
            "default_unit": clean_str(row.get("unidad")) or "kg",
        }
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        st, resp = api.post("/api/v1/materials/", payload)
        if st in (200, 201):
            mapping[codigo] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): {st} - {resp}")
    if not dry_run and stats.skipped:
        st, resp = api.get("/api/v1/materials/", {"limit": 500})
        if st == 200:
            for item in resp.get("items", []):
                if item.get("code") and item["code"] not in mapping:
                    mapping[item["code"]] = item["id"]
    return stats, mapping


# ---------------------------------------------------------------------------
# Phase: Precios
# ---------------------------------------------------------------------------

def load_prices(
    api: APIClient, rows: list[dict], mat_map: dict, dry_run: bool
) -> Stats:
    """Para cada fila: hace 2 POSTs separados a /api/v1/price-lists/ (uno por
    price_type=purchase, otro por price_type=sale). Pasa 'notas' a AMBOS POSTs.
    Skip cuando ambos precios son 0/vacios/blank.
    """
    stats = Stats("Precios")
    for row in rows:
        codigo = clean_str(row.get("material_codigo"))
        if not codigo:
            stats.add_error(f"  Fila {row['_row']}: material_codigo obligatorio")
            continue
        mat_id = find_in_map_ci(mat_map, codigo)
        if not mat_id:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): material no encontrado")
            continue
        precio_compra = parse_decimal(row.get("precio_compra"), 0)
        precio_venta = parse_decimal(row.get("precio_venta"), 0)
        if precio_compra <= 0 and precio_venta <= 0:
            stats.skipped += 1
            continue
        notas = clean_str(row.get("notas"))

        for price_type, value in (("purchase", precio_compra), ("sale", precio_venta)):
            if value <= 0:
                continue
            if dry_run:
                stats.created += 1
                continue
            payload: dict = {
                "material_id": mat_id,
                "price_type": price_type,
                "price": float(value),
            }
            if notas:
                payload["notes"] = notas  # fix: pasar notas a AMBOS POSTs
            st, resp = api.post("/api/v1/price-lists/", payload)
            if st in (200, 201):
                stats.created += 1
            elif is_duplicate(st, resp):
                stats.skipped += 1
            else:
                stats.add_error(
                    f"  Fila {row['_row']} ({codigo}, {price_type}): {st} - {resp}"
                )
    return stats


# ---------------------------------------------------------------------------
# Phase: Terceros
# ---------------------------------------------------------------------------

def load_third_parties(
    api: APIClient, rows: list[dict], tp_cat_map: dict, dry_run: bool
) -> tuple[Stats, dict]:
    stats = Stats("Terceros")
    mapping: dict[str, str] = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        cats_raw = clean_str(row.get("categorias"))
        if not cats_raw:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): categorias obligatorias")
            continue
        category_ids = []
        cat_errors = []
        for cat_name in [c.strip() for c in cats_raw.split(",") if c.strip()]:
            cat_id = find_in_map_ci(tp_cat_map, cat_name)
            if cat_id:
                category_ids.append(cat_id)
            else:
                cat_errors.append(cat_name)
        if cat_errors:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): categorias no encontradas: {', '.join(cat_errors)}"
            )
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-tp-{nombre}"
            continue
        payload: dict = {
            "name": nombre,
            "initial_balance": float(parse_decimal(row.get("saldo_inicial"), 0)),
            "category_ids": category_ids,
        }
        for src_field, dst_field in [
            ("identificacion", "identification_number"),
            ("email", "email"),
            ("telefono", "phone"),
            ("direccion", "address"),
        ]:
            v = clean_str(row.get(src_field))
            if v:
                payload[dst_field] = str(v)
        provision_type = clean_str(row.get("provision_type"))
        if provision_type:
            payload["provision_type"] = provision_type
        st, resp = api.post("/api/v1/third-parties/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        st, resp = api.get("/api/v1/third-parties/", {"limit": 5000})
        if st == 200:
            for item in resp.get("items", []):
                if item.get("name") and item["name"] not in mapping:
                    mapping[item["name"]] = item["id"]
    return stats, mapping


# ---------------------------------------------------------------------------
# Phase: Cuentas
# ---------------------------------------------------------------------------

def load_money_accounts(api: APIClient, rows: list[dict], dry_run: bool) -> tuple[Stats, dict]:
    stats = Stats("Cuentas")
    mapping: dict[str, str] = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        tipo_raw = clean_str(row.get("tipo"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        if not tipo_raw:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): tipo obligatorio")
            continue
        tipo = ACCOUNT_TYPE_MAP.get(tipo_raw.lower().strip())
        if not tipo:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): tipo '{tipo_raw}' invalido")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-acc-{nombre}"
            continue
        payload: dict = {
            "name": nombre,
            "account_type": tipo,
            "initial_balance": float(parse_decimal(row.get("saldo_inicial"), 0)),
        }
        num = clean_str(row.get("numero_cuenta"))
        if num:
            payload["account_number"] = str(num)
        banco = clean_str(row.get("banco"))
        if banco:
            payload["bank_name"] = banco
        st, resp = api.post("/api/v1/money-accounts/", payload)
        if st in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    if not dry_run and stats.skipped:
        fill_existing(api, "/api/v1/money-accounts/", mapping)
    return stats, mapping


# ---------------------------------------------------------------------------
# Phase: Inventario
# ---------------------------------------------------------------------------

def seed_inventory(
    api: APIClient,
    rows: list[dict],
    mat_map: dict,
    wh_map: dict,
    inventory_reason: str,
    default_warehouse: str,
    dry_run: bool,
) -> Stats:
    stats = Stats("Inventario")
    today = date.today().isoformat()
    for row in rows:
        codigo = clean_str(row.get("material_codigo"))
        if not codigo:
            stats.add_error(f"  Fila {row['_row']}: material_codigo obligatorio")
            continue
        mat_id = find_in_map_ci(mat_map, codigo)
        if not mat_id:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): material no encontrado")
            continue
        bodega_nombre = clean_str(row.get("bodega")) or default_warehouse
        wh_id = find_in_map_ci(wh_map, bodega_nombre)
        if not wh_id:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): bodega '{bodega_nombre}' no encontrada")
            continue
        cantidad = parse_decimal(row.get("cantidad"), 0)
        costo = parse_decimal(row.get("costo_unitario"), 0)
        if cantidad <= 0:
            stats.skipped += 1
            continue
        if costo <= 0:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): costo_unitario debe ser > 0")
            continue
        fecha = parse_date(row.get("fecha")) or today
        if dry_run:
            stats.created += 1
            continue
        payload = {
            "material_id": mat_id,
            "warehouse_id": wh_id,
            "quantity": float(cantidad),
            "unit_cost": float(costo),
            "date": fecha,
            "reason": inventory_reason,
        }
        st, resp = api.post("/api/v1/inventory/adjustments/increase", payload)
        if st in (200, 201):
            stats.created += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({codigo}): {st} - {resp}")
    return stats


# ---------------------------------------------------------------------------
# Phase: Activos Fijos
# ---------------------------------------------------------------------------

def seed_fixed_assets(
    api: APIClient,
    rows: list[dict],
    tp_map: dict,
    acc_map: dict,
    ec_map: dict,
    bu_map: dict,
    dry_run: bool,
) -> Stats:
    """Carga activos fijos. Diferencias vs script anterior:
      - REMOVIDA columna 'tasa_depreciacion' (era buggy: vida_util_meses es la fuente unica).
      - NUEVA columna 'unidades_negocio' (CSV) → applicable_business_unit_ids (lista JSONB).
      - Validacion client-side en dry-run:
            depreciacion_acumulada <= valor_compra - valor_residual.
      - historical_load=True si proveedor Y cuenta_pago vacios (no MoneyMovement,
        no balances; el "pago" ocurrio en sistema origen).
    """
    stats = Stats("ActivosFijos")
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.add_error(f"  Fila {row['_row']}: nombre obligatorio")
            continue
        valor_compra = parse_decimal(row.get("valor_compra"), 0)
        if valor_compra <= 0:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): valor_compra debe ser > 0")
            continue
        valor_residual = parse_decimal(row.get("valor_residual"), 0)
        vida_meses = parse_int(row.get("vida_util_meses"), 0)
        if vida_meses <= 0:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): vida_util_meses debe ser > 0")
            continue
        rate = (Decimal("100") / Decimal(vida_meses)).quantize(Decimal("0.0001"))
        if rate < Decimal("0.01") or rate > Decimal("100"):
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): tasa derivada {rate} fuera de [0.01, 100]"
            )
            continue
        fecha_compra = parse_date(row.get("fecha_compra"))
        fecha_inicio = parse_date(row.get("fecha_inicio_depreciacion")) or fecha_compra
        if not fecha_compra:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): fecha_compra obligatoria")
            continue

        dep_acumulada = parse_decimal(row.get("depreciacion_acumulada"), 0)

        # Validacion client-side: depreciacion_acumulada <= valor_compra - valor_residual
        depreciable_max = valor_compra - valor_residual
        if dep_acumulada > depreciable_max:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): depreciacion_acumulada ({dep_acumulada}) "
                f"excede valor_compra - valor_residual ({depreciable_max}). "
                f"Maximo depreciable es la base depreciable."
            )
            continue
        if dep_acumulada < 0:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): depreciacion_acumulada no puede ser negativa"
            )
            continue

        cat_nombre = clean_str(row.get("categoria_gasto"))
        cat_id = find_in_map_ci(ec_map, cat_nombre) if cat_nombre else None
        if not cat_id:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): categoria_gasto '{cat_nombre}' no encontrada"
            )
            continue

        proveedor = clean_str(row.get("proveedor"))
        cuenta = clean_str(row.get("cuenta_pago"))
        supplier_id = find_in_map_ci(tp_map, proveedor) if proveedor else None
        source_account_id = find_in_map_ci(acc_map, cuenta) if cuenta else None
        if proveedor and not supplier_id:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): proveedor '{proveedor}' no encontrado")
            continue
        if cuenta and not source_account_id:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): cuenta '{cuenta}' no encontrada")
            continue

        # UN single
        un_nombre = clean_str(row.get("unidad_negocio"))
        bu_id = find_in_map_ci(bu_map, un_nombre) if un_nombre else None
        if un_nombre and not bu_id:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): UN '{un_nombre}' no encontrada")
            continue

        # UN list (CSV) → applicable_business_unit_ids
        un_csv = parse_csv(row.get("unidades_negocio"))
        applicable_bu_ids: list[str] = []
        bu_errors: list[str] = []
        for un_name in un_csv:
            uid = find_in_map_ci(bu_map, un_name)
            if uid:
                applicable_bu_ids.append(uid)
            else:
                bu_errors.append(un_name)
        if bu_errors and not dry_run:
            stats.add_error(
                f"  Fila {row['_row']} ({nombre}): unidades_negocio no encontradas: "
                f"{', '.join(bu_errors)}"
            )
            continue

        # historical_load: si no hay supplier ni cuenta, asumir carga historica
        historical = not supplier_id and not source_account_id
        if dry_run:
            kind = "historical" if historical else ("supplier" if supplier_id else "account")
            print(
                f"    (dry-run) {nombre}: {kind}, valor={valor_compra}, "
                f"dep_acum={dep_acumulada}, bus_units={len(applicable_bu_ids)}"
            )
            stats.created += 1
            continue
        payload: dict = {
            "name": nombre,
            "purchase_date": fecha_compra,
            "purchase_value": float(valor_compra),
            "salvage_value": float(valor_residual),
            "depreciation_rate": float(rate),
            "depreciation_start_date": fecha_inicio,
            "expense_category_id": cat_id,
            "accumulated_depreciation": float(dep_acumulada),
            "historical_load": historical,
        }
        codigo = clean_str(row.get("codigo_activo"))
        if codigo:
            payload["asset_code"] = codigo
        notas = clean_str(row.get("notas"))
        if notas:
            payload["notes"] = notas
        if supplier_id:
            payload["supplier_id"] = supplier_id
        if source_account_id:
            payload["source_account_id"] = source_account_id
        if bu_id:
            payload["business_unit_id"] = bu_id
        if applicable_bu_ids:
            payload["applicable_business_unit_ids"] = applicable_bu_ids
        st, resp = api.post("/api/v1/fixed-assets/", payload)
        if st in (200, 201):
            stats.created += 1
        elif is_duplicate(st, resp):
            stats.skipped += 1
        else:
            stats.add_error(f"  Fila {row['_row']} ({nombre}): {st} - {resp}")
    return stats


# ---------------------------------------------------------------------------
# Phase: Verification
# ---------------------------------------------------------------------------

def run_verification(api: APIClient, expected: dict, tolerance: Decimal) -> tuple[int, int]:
    """Retorna (passed, failed). Imprime detalle de cada check."""
    print("\n" + "=" * 60)
    print("  VERIFICACION POST-MIGRACION")
    print("=" * 60)
    passed = 0
    failed = 0

    def _check_exact(label, actual, expected_value):
        nonlocal passed, failed
        if actual == expected_value:
            print(f"  OK    {label}: {actual}")
            passed += 1
        else:
            print(f"  FAIL  {label}: esperado={expected_value}, actual={actual}")
            failed += 1

    def _check_at_least(label, actual, minimum):
        nonlocal passed, failed
        if actual >= minimum:
            print(f"  OK    {label}: {actual} (>= {minimum})")
            passed += 1
        else:
            print(f"  FAIL  {label}: actual={actual} < esperado_minimo={minimum}")
            failed += 1

    def _check_tolerance(label, actual, expected_value):
        nonlocal passed, failed
        actual_d = Decimal(str(actual))
        expected_d = Decimal(str(expected_value))
        diff = abs(actual_d - expected_d)
        if diff <= tolerance:
            note = f" (diff={diff})" if diff > 0 else ""
            print(f"  OK    {label}: {actual_d}{note}")
            passed += 1
        else:
            print(
                f"  FAIL  {label}: esperado={expected_d}, actual={actual_d}, "
                f"diff={diff} > tolerancia={tolerance}"
            )
            failed += 1

    # # business units
    st, resp = api.get("/api/v1/business-units/", {"limit": 1})
    if st == 200:
        actual = resp.get("total", len(resp.get("items", []))) if isinstance(resp, dict) else len(resp)
        _check_at_least("# unidades de negocio", actual, expected["business_units_count"])
    else:
        print(f"  ERR   business-units: {st} - {resp}")
        failed += 1

    # # warehouses
    st, resp = api.get("/api/v1/warehouses/", {"limit": 1})
    if st == 200:
        actual = resp.get("total", len(resp.get("items", []))) if isinstance(resp, dict) else len(resp)
        _check_at_least("# bodegas", actual, expected["warehouses_count"])
    else:
        print(f"  ERR   warehouses: {st} - {resp}")
        failed += 1

    # # materiales activos
    st, resp = api.get("/api/v1/materials/", {"is_active": "true", "limit": 1})
    if st == 200:
        _check_at_least("# materiales activos", resp.get("total", 0), expected["materials_count"])
    else:
        print(f"  ERR   materiales: {st} - {resp}")
        failed += 1

    # # terceros activos
    st, resp = api.get("/api/v1/third-parties/", {"is_active": "true", "limit": 1})
    if st == 200:
        _check_at_least("# terceros activos", resp.get("total", 0), expected["third_parties_count"])
    else:
        print(f"  ERR   terceros: {st} - {resp}")
        failed += 1

    # # precios (lista de precios)
    st, resp = api.get("/api/v1/price-lists/", {"limit": 1})
    if st == 200:
        actual = resp.get("total", len(resp.get("items", []))) if isinstance(resp, dict) else 0
        _check_at_least("# precios cargados", actual, expected["prices_count"])
    else:
        # endpoint puede no soportar total: degradar a soft-check
        print(f"  WARN  price-lists no devolvio listado: {st} - {resp}")

    # Sum cuentas (treasury-dashboard)
    st, resp = api.get("/api/v1/reports/treasury-dashboard")
    if st == 200:
        _check_tolerance("Sum cuentas operativas", resp.get("total_all_accounts", 0), expected["accounts_sum"])
    else:
        print(f"  ERR   treasury-dashboard: {st} - {resp}")
        failed += 1

    # Activos fijos
    st, resp = api.get("/api/v1/fixed-assets/", {"limit": 200})
    if st == 200:
        items = resp.get("items", [])
        total_cv = sum(Decimal(str(i.get("current_value", 0))) for i in items)
        _check_exact("# activos fijos", len(items), expected["fixed_assets_count"])
        _check_tolerance("Sum current_value activos", total_cv, expected["fixed_assets_current_value_sum"])
    else:
        print(f"  ERR   fixed-assets: {st} - {resp}")
        failed += 1

    # Balance Detallado cuadra
    st, resp = api.get("/api/v1/reports/balance-detailed")
    if st == 200:
        ver = resp.get("verification", {})
        is_bal = ver.get("is_balanced", False)
        formula_result = ver.get("result", 0)
        if is_bal or abs(Decimal(str(formula_result))) <= tolerance:
            print(
                f"  OK    Balance General cuadra: total_assets={resp.get('total_assets')} "
                f"= total_liabilities + equity (diff={formula_result})"
            )
            passed += 1
        else:
            print(
                f"  FAIL  Balance General NO cuadra: total_assets={resp.get('total_assets')}, "
                f"total_liabilities={resp.get('total_liabilities')}, "
                f"equity={resp.get('equity')}, accumulated_profit={resp.get('accumulated_profit')}, "
                f"distributed_profit={resp.get('distributed_profit')}, diff={formula_result}"
            )
            failed += 1
    else:
        print(f"  ERR   balance-detailed: {st} - {resp}")
        failed += 1

    return passed, failed


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Migracion generica de organizaciones a EcoBalance",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--file", required=True, help="Ruta al Excel de migracion")
    parser.add_argument("--api-url", default="http://localhost:8000")
    parser.add_argument(
        "--superuser-email",
        required=False,
        default=None,
        help="Email del superuser. Obligatorio con --apply, opcional con --dry-run.",
    )
    parser.add_argument("--superuser-password", required=False, default=None)
    parser.add_argument(
        "--org-name",
        required=True,
        help="Nombre de la organizacion a crear (ej: 'Mi Empresa SAS')",
    )
    parser.add_argument(
        "--admin-email", required=True, help="Email del admin de la nueva org"
    )
    parser.add_argument("--admin-name", required=True, help="Nombre completo del admin")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=True,
        help="Default: dry-run (validacion sin escribir)",
    )
    parser.add_argument(
        "--apply", dest="dry_run", action="store_false", help="Ejecutar de verdad"
    )
    parser.add_argument(
        "--reset-org",
        action="store_true",
        help="Soft-delete org existente antes (PRE GO-LIVE)",
    )
    parser.add_argument(
        "--balance-tolerance",
        default="100.00",
        help="Tolerancia COP para checks (default 100.00)",
    )
    parser.add_argument(
        "--skip-phase",
        default="",
        help="Lista CSV de fases a saltar (ej: '11,12' para no resemilla inventario y activos)",
    )
    args = parser.parse_args()

    dry_run = args.dry_run
    skip_phases = {int(p.strip()) for p in args.skip_phase.split(",") if p.strip()}
    tolerance = parse_decimal(args.balance_tolerance, "100.00")
    inventory_reason = f"Carga inicial migracion {args.org_name}"

    print("\n" + "=" * 60)
    print(f"  MIGRACION → '{args.org_name}'")
    print(f"  {'DRY-RUN (validacion)' if dry_run else 'APPLY (escritura)'}")
    print("=" * 60)
    print(f"  Excel:           {args.file}")
    print(f"  API:             {args.api_url}")
    print(f"  Admin:           {args.admin_email} ({args.admin_name})")
    print(f"  Tolerancia:      {tolerance} COP")
    print(f"  Inventory reason: '{inventory_reason}'")
    if skip_phases:
        print(f"  Skip:            fases {sorted(skip_phases)}")
    if args.reset_org:
        print("  RESET-ORG:       habilitado")
    print()

    # Cargar Excel
    try:
        wb = load_workbook(args.file, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: archivo '{args.file}' no encontrado")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR leyendo Excel: {e}")
        sys.exit(1)

    print(f"Hojas: {', '.join(wb.sheetnames)}\n")

    # Login (solo si --apply)
    api = APIClient(args.api_url)
    if not dry_run:
        if not args.superuser_email or not args.superuser_password:
            print("ERROR: --apply requiere --superuser-email y --superuser-password")
            sys.exit(1)
        print(f"[*] Autenticando como superuser '{args.superuser_email}'...")
        if not api.login(args.superuser_email, args.superuser_password):
            sys.exit(1)
        print("    OK\n")
    else:
        print("[*] Dry-run: saltando autenticacion (validacion offline del Excel)\n")

    # Phase 0: reset si pidieron
    if args.reset_org:
        print("[*] --reset-org: soft-delete org existente")
        if not dry_run and not reset_org(api, args.org_name, dry_run):
            sys.exit(1)
        if dry_run:
            print(f"  (dry-run, --reset-org) Soft-deletaria org '{args.org_name}' si existiera")
        print()

    # Phase 1: org + admin
    print("[1] Crear/detectar org + admin")
    if dry_run:
        print(f"  (dry-run) Crearia (o reusaria) org '{args.org_name}' con admin {args.admin_email}")
        org_id = "00000000-0000-0000-0000-000000000000"
    else:
        org_id = ensure_org_and_admin(api, args.org_name, args.admin_email, args.admin_name, dry_run)
        if not org_id:
            sys.exit(1)
    api.set_org(org_id)
    print()

    all_stats: list[Stats] = []

    def _skip_or_fill(map_var: dict, path: str, key: str = "name") -> dict:
        """En skip de fase: si --apply, recarga IDs reales desde API; en dry-run, deja vacio."""
        if not dry_run:
            fill_existing(api, path, map_var, key)
        return map_var

    # Phase 2: UnidadesNegocio
    if 2 not in skip_phases:
        print("[2] UnidadesNegocio")
        rows = read_sheet(wb, "UnidadesNegocio")
        s, bu_map = load_business_units(api, rows, dry_run)
        all_stats.append(s); print(s.report())
    else:
        bu_map = _skip_or_fill({}, "/api/v1/business-units/")

    # Phase 3: Bodegas
    if 3 not in skip_phases:
        print("[3] Bodegas")
        rows = read_sheet(wb, "Bodegas")
        s, wh_map = load_warehouses(api, rows, dry_run)
        all_stats.append(s); print(s.report())
    else:
        wh_map = _skip_or_fill({}, "/api/v1/warehouses/")

    # Phase 4: CategoriaMateriales
    if 4 not in skip_phases:
        print("[4] CategoriaMateriales")
        rows = read_sheet(wb, "CategoriaMateriales")
        s, cat_map = load_material_categories(api, rows, dry_run)
        all_stats.append(s); print(s.report())
    else:
        cat_map = _skip_or_fill({}, "/api/v1/material-categories/")

    # Phase 5: CategoriaGastos
    if 5 not in skip_phases:
        print("[5] CategoriaGastos")
        rows = read_sheet(wb, "CategoriaGastos")
        s, ec_map = load_expense_categories(api, rows, bu_map, dry_run)
        all_stats.append(s); print(s.report())
    else:
        ec_map = {}
        if not dry_run:
            st, resp = api.get("/api/v1/expense-categories/", {"limit": 500})
            if st == 200:
                for it in resp.get("items", []):
                    ec_map[it["name"]] = it["id"]

    # Phase 6: CategoriaTerceros (extras only; las 7 default vienen con la org)
    if 6 not in skip_phases:
        print("[6] CategoriaTerceros (extras; las 7 default ya estan en la org)")
        # Primero traer defaults via API; en dry-run usar placeholders
        defaults_map = fetch_default_tp_categories(api, dry_run)
        rows = read_sheet(wb, "CategoriaTerceros")
        s, tp_cat_map = load_third_party_categories_extras(api, rows, defaults_map, dry_run)
        all_stats.append(s); print(s.report())
    else:
        tp_cat_map = _skip_or_fill({}, "/api/v1/third-party-categories/")

    # Phase 7: Materiales (depende de MapeoUN como fallback)
    print("[*] MapeoUN (referencia, no escribe)")
    rows = read_sheet(wb, "MapeoUN")
    un_by_category = build_un_map(rows)
    print(f"  [MapeoUN] {len(un_by_category)} categorias mapeadas a UN")
    for cat, un in sorted(un_by_category.items()):
        print(f"    {cat} → {un}")

    if 7 not in skip_phases:
        print("[7] Materiales")
        rows = read_sheet(wb, "Materiales")
        s, mat_map = load_materials(api, rows, cat_map, bu_map, un_by_category, dry_run)
        all_stats.append(s); print(s.report())
    else:
        mat_map = {}
        if not dry_run:
            st, resp = api.get("/api/v1/materials/", {"limit": 500})
            if st == 200:
                for it in resp.get("items", []):
                    if it.get("code"):
                        mat_map[it["code"]] = it["id"]

    # Phase 8: Precios
    if 8 not in skip_phases:
        print("[8] Precios")
        rows = read_sheet(wb, "Precios")
        s = load_prices(api, rows, mat_map, dry_run)
        all_stats.append(s); print(s.report())

    # Phase 9: Terceros
    if 9 not in skip_phases:
        print("[9] Terceros")
        rows = read_sheet(wb, "Terceros")
        s, tp_map = load_third_parties(api, rows, tp_cat_map, dry_run)
        all_stats.append(s); print(s.report())
    else:
        tp_map = {}
        if not dry_run:
            st, resp = api.get("/api/v1/third-parties/", {"limit": 5000})
            if st == 200:
                for it in resp.get("items", []):
                    tp_map[it["name"]] = it["id"]

    # Phase 10: Cuentas
    if 10 not in skip_phases:
        print("[10] Cuentas")
        rows = read_sheet(wb, "Cuentas")
        s, acc_map = load_money_accounts(api, rows, dry_run)
        all_stats.append(s); print(s.report())
    else:
        acc_map = _skip_or_fill({}, "/api/v1/money-accounts/")

    # Default warehouse para inventario (primer key del mapping, suele ser "Principal")
    default_warehouse_name = next(iter(wh_map.keys()), DEFAULT_WAREHOUSE_NAME)

    # Phase 11: Inventario
    if 11 not in skip_phases:
        print("[11] Inventario")
        rows = read_sheet(wb, "Inventario")
        s = seed_inventory(
            api, rows, mat_map, wh_map, inventory_reason, default_warehouse_name, dry_run
        )
        all_stats.append(s); print(s.report())

    # Phase 12: ActivosFijos
    if 12 not in skip_phases:
        print("[12] ActivosFijos")
        rows = read_sheet(wb, "ActivosFijos")
        s = seed_fixed_assets(api, rows, tp_map, acc_map, ec_map, bu_map, dry_run)
        all_stats.append(s); print(s.report())

    wb.close()

    # Resumen
    print("\n" + "=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    total_created = sum(s.created for s in all_stats)
    total_skipped = sum(s.skipped for s in all_stats)
    total_errors = sum(s.errors for s in all_stats)
    print(f"  Creados:  {total_created}")
    print(f"  Omitidos: {total_skipped}")
    print(f"  Errores:  {total_errors}")

    all_errors: list[str] = []
    for s in all_stats:
        all_errors.extend(s.error_details)
    if all_errors:
        print("\n  DETALLE ERRORES:")
        for e in all_errors:
            print(f"    {e}")

    if total_errors > 0:
        print("\nABORTANDO: hubo errores arriba.")
        sys.exit(1)

    # Verificacion (solo si --apply y todas las fases corrieron)
    if not dry_run and not skip_phases:
        # Calcular esperados leyendo el Excel
        wb2 = load_workbook(args.file, read_only=True, data_only=True)
        bu_rows = read_sheet(wb2, "UnidadesNegocio")
        wh_rows = read_sheet(wb2, "Bodegas")
        mat_rows = read_sheet(wb2, "Materiales")
        price_rows = read_sheet(wb2, "Precios")
        tp_rows = read_sheet(wb2, "Terceros")
        acc_rows = read_sheet(wb2, "Cuentas")
        fa_rows = read_sheet(wb2, "ActivosFijos")
        wb2.close()

        # # de precios esperados = cantidad de POSTs (1 por precio_compra>0 + 1 por precio_venta>0)
        prices_count_expected = 0
        for r in price_rows:
            if parse_decimal(r.get("precio_compra"), 0) > 0:
                prices_count_expected += 1
            if parse_decimal(r.get("precio_venta"), 0) > 0:
                prices_count_expected += 1

        expected = {
            "business_units_count": len(bu_rows),
            "warehouses_count": max(len(wh_rows), 1),  # min 1 (fallback)
            "materials_count": len(mat_rows),
            "third_parties_count": len(tp_rows),
            "prices_count": prices_count_expected,
            "accounts_sum": sum(parse_decimal(r.get("saldo_inicial"), 0) for r in acc_rows),
            "fixed_assets_count": len(fa_rows),
            "fixed_assets_current_value_sum": sum(
                parse_decimal(r.get("valor_compra"), 0) - parse_decimal(r.get("depreciacion_acumulada"), 0)
                for r in fa_rows
            ),
        }
        passed, failed = run_verification(api, expected, tolerance)
        print(f"\n  Verificacion: {passed} OK, {failed} FAIL")
        if failed > 0:
            sys.exit(1)

    print("\nDone.\n")


if __name__ == "__main__":
    main()
