"""
Carga inicial de datos maestros desde archivo Excel.

Lee un archivo Excel con 10 hojas (en orden de dependencia) y carga los datos
via API REST autenticada. Permite re-ejecucion segura (omite duplicados).

Hojas del Excel (en orden):
  1. UnidadesNegocio      → Unidades de negocio
  2. CategoriaMateriales  → Categorías de materiales
  3. Bodegas              → Bodegas / almacenes
  4. Cuentas              → Cuentas monetarias
  5. CategoriaGastos      → Categorías de gastos (con jerarquia padre-hijo + UN)
  6. CategoriaTerceros    → Categorías de terceros (con behavior_type + jerarquia)
  7. Terceros             → Proveedores, clientes, etc. (asignados a categorias por nombre)
  8. Materiales           → Materiales (requiere CategoriaMateriales + UnidadesNegocio)
  9. Precios              → Lista de precios (requiere Materiales)
  10. Roles               → Roles personalizados (opcional)

Uso:
    cd backend
    ./venv/bin/python scripts/load_initial_data.py \\
        --file docs/CargaInicial_EcoBalance_v2_cliente.xlsx \\
        --email admin@empresa.com \\
        --password secreto \\
        --org-id 550e8400-e29b-41d4-a716-446655440000

    # Solo validar sin crear:
    ./venv/bin/python scripts/load_initial_data.py --file datos.xlsx --dry-run \\
        --email admin@test.com --password test --org-id <uuid>
"""
import argparse
import sys
from decimal import Decimal, InvalidOperation

import requests

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no esta instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

ACCOUNT_TYPE_MAP = {
    "efectivo": "cash",
    "banco": "bank",
    "digital": "digital",
}

BEHAVIOR_TYPE_MAP = {
    "proveedor_material": "material_supplier",
    "proveedor_servicios": "service_provider",
    "cliente": "customer",
    "inversionista": "investor",
    "generico": "generic",
    "provision": "provision",
    "pasivo": "liability",
    # Aceptar tambien valores backend directos
    "material_supplier": "material_supplier",
    "service_provider": "service_provider",
    "customer": "customer",
    "investor": "investor",
    "generic": "generic",
    "liability": "liability",
}


def parse_bool(value) -> bool:
    if value is None:
        return False
    return str(value).strip().upper() in ("SI", "SÍ", "YES", "1", "TRUE")


def parse_decimal(value, default=0) -> float:
    if value is None:
        return default
    s = str(value).strip()
    if s in ("", "$", "$ -", "-"):
        return default
    try:
        return float(Decimal(s.replace("$", "").replace(",", "").strip()))
    except (InvalidOperation, ValueError):
        return default


def clean_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def read_sheet(wb, sheet_name: str) -> list[dict]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [clean_str(h) or f"col_{i}" for i, h in enumerate(rows[0])]
    headers = [h.lower().strip().rstrip("*").strip() for h in headers]

    result = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if all(v is None for v in row):
            continue
        data = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                data[headers[col_idx]] = val
        data["_row"] = row_idx
        result.append(data)
    return result


def find_in_map_ci(mapping: dict, key: str) -> str | None:
    """Buscar en mapping case-insensitive. Retorna el ID o None."""
    if not key:
        return None
    key_lower = key.lower().strip()
    for k, v in mapping.items():
        if k.lower().strip() == key_lower:
            return v
    return None


# ---------------------------------------------------------------------------
# API Client
# ---------------------------------------------------------------------------

class APIClient:
    def __init__(self, base_url: str, org_id: str):
        self.base_url = base_url.rstrip("/")
        self.org_id = org_id
        self.token = None
        self.session = requests.Session()

    def login(self, email: str, password: str) -> bool:
        resp = self.session.post(
            f"{self.base_url}/api/v1/auth/login/json",
            json={"email": email, "password": password},
        )
        if resp.status_code == 200:
            self.token = resp.json()["access_token"]
            self.session.headers.update({
                "Authorization": f"Bearer {self.token}",
                "X-Organization-ID": self.org_id,
            })
            return True
        print(f"  ERROR login: {resp.status_code} - {resp.text}")
        return False

    def post(self, path: str, data: dict) -> tuple[int, dict | str]:
        resp = self.session.post(f"{self.base_url}{path}", json=data)
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, resp.text

    def get(self, path: str, params: dict | None = None) -> tuple[int, dict | str]:
        resp = self.session.get(f"{self.base_url}{path}", params=params)
        try:
            return resp.status_code, resp.json()
        except Exception:
            return resp.status_code, resp.text


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

class Stats:
    def __init__(self):
        self.created = 0
        self.skipped = 0
        self.errors = 0
        self.error_details: list[str] = []

    def report(self, sheet: str) -> str:
        parts = []
        if self.created:
            parts.append(f"{self.created} creados")
        if self.skipped:
            parts.append(f"{self.skipped} omitidos")
        if self.errors:
            parts.append(f"{self.errors} errores")
        return f"  {sheet}: {', '.join(parts) if parts else 'vacio'}"


def _is_duplicate(status_code, resp):
    """Detectar si el error es un duplicado (ya existe)."""
    if status_code == 409:
        return True
    resp_str = str(resp).lower()
    return status_code == 400 and ("ya existe" in resp_str or "duplicate" in resp_str)


def _fill_mapping_from_api(api, path, rows, mapping, key_field="nombre", resp_key="name"):
    """Cargar IDs de entidades existentes para completar el mapping."""
    st, resp = api.get(path, {"limit": 1000})
    if st == 200:
        items = resp.get("items", resp) if isinstance(resp, dict) else resp
        if isinstance(items, dict):
            items = items.get("items", [])
        existing = {item[resp_key]: item["id"] for item in items if resp_key in item}
        for row in rows:
            nombre = clean_str(row.get(key_field))
            if nombre and nombre not in mapping:
                # Case-insensitive match
                for k, v in existing.items():
                    if k.lower().strip() == nombre.lower().strip():
                        mapping[nombre] = v
                        break


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_unidades_negocio(api, rows, dry_run):
    stats = Stats()
    mapping = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue
        payload = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        status, resp = api.post("/api/v1/business-units/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")
    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/business-units/", rows, mapping)
    return stats, mapping


def load_categorias_material(api, rows, dry_run):
    stats = Stats()
    mapping = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue
        payload = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        status, resp = api.post("/api/v1/material-categories/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")
    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/material-categories/", rows, mapping)
    return stats, mapping


def load_bodegas(api, rows, dry_run):
    stats = Stats()
    mapping = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue
        payload = {"name": nombre}
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc
        addr = clean_str(row.get("direccion"))
        if addr:
            payload["address"] = addr
        status, resp = api.post("/api/v1/warehouses/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")
    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/warehouses/", rows, mapping)
    return stats, mapping


def load_cuentas(api, rows, dry_run):
    stats = Stats()
    mapping = {}
    for row in rows:
        nombre = clean_str(row.get("nombre"))
        tipo_raw = clean_str(row.get("tipo"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue
        if not tipo_raw:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): tipo es obligatorio")
            continue
        tipo = ACCOUNT_TYPE_MAP.get(tipo_raw.lower())
        if not tipo:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): tipo '{tipo_raw}' invalido")
            continue
        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue
        payload = {
            "name": nombre,
            "account_type": tipo,
            "initial_balance": parse_decimal(row.get("saldo_inicial"), 0),
        }
        num = clean_str(row.get("numero_cuenta"))
        if num:
            payload["account_number"] = num
        banco = clean_str(row.get("banco"))
        if banco:
            payload["bank_name"] = banco
        status, resp = api.post("/api/v1/money-accounts/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")
    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/money-accounts/", rows, mapping)
    return stats, mapping


def load_categoria_gastos(api, rows, un_map, dry_run):
    """Carga categorias de gasto con jerarquia padre-hijo + asignacion UN.
    Procesa en 2 pasadas: primero raices (sin padre), luego hijas."""
    stats = Stats()
    mapping = {}  # nombre → id

    # Separar raices e hijas
    raices = [r for r in rows if not clean_str(r.get("padre"))]
    hijas = [r for r in rows if clean_str(r.get("padre"))]

    for row in raices + hijas:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue

        padre_nombre = clean_str(row.get("padre"))
        parent_id = None
        if padre_nombre:
            parent_id = find_in_map_ci(mapping, padre_nombre)
            if not parent_id and not dry_run:
                stats.errors += 1
                stats.error_details.append(f"  Fila {row['_row']} ({nombre}): padre '{padre_nombre}' no encontrado")
                continue

        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue

        payload = {
            "name": nombre,
            "is_direct_expense": parse_bool(row.get("es_directo")) if not padre_nombre else False,
        }
        if parent_id:
            payload["parent_id"] = parent_id

        # Asignacion UN
        asignacion = clean_str(row.get("asignacion_un"))
        if asignacion:
            asignacion_lower = asignacion.lower().strip()
            if asignacion_lower == "directo":
                un_nombre = clean_str(row.get("unidad_negocio"))
                un_id = find_in_map_ci(un_map, un_nombre) if un_nombre else None
                if un_id:
                    payload["default_business_unit_id"] = un_id
            elif asignacion_lower == "compartido":
                uns_raw = clean_str(row.get("unidades_negocio"))
                if uns_raw:
                    un_ids = []
                    for un_name in [u.strip() for u in uns_raw.split(",") if u.strip()]:
                        uid = find_in_map_ci(un_map, un_name)
                        if uid:
                            un_ids.append(uid)
                    if un_ids:
                        payload["default_applicable_business_unit_ids"] = un_ids

        status, resp = api.post("/api/v1/expense-categories/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")

    if stats.skipped > 0 and not dry_run:
        st, resp = api.get("/api/v1/expense-categories/", {"limit": 500})
        if st == 200:
            for item in resp.get("items", []):
                if item["name"] not in mapping:
                    mapping[item["name"]] = item["id"]

    return stats, mapping


def load_categoria_terceros(api, rows, dry_run):
    """Carga categorias de terceros con behavior_type + jerarquia.
    Procesa en 2 pasadas: primero raices, luego hijas."""
    stats = Stats()
    mapping = {}  # nombre → id

    raices = [r for r in rows if not clean_str(r.get("padre"))]
    hijas = [r for r in rows if clean_str(r.get("padre"))]

    for row in raices + hijas:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue

        padre_nombre = clean_str(row.get("padre"))
        parent_id = None
        if padre_nombre:
            parent_id = find_in_map_ci(mapping, padre_nombre)
            if not parent_id and not dry_run:
                stats.errors += 1
                stats.error_details.append(f"  Fila {row['_row']} ({nombre}): padre '{padre_nombre}' no encontrado")
                continue

        # behavior_type: obligatorio para raices, opcional para hijas
        bt_raw = clean_str(row.get("tipo_comportamiento"))
        behavior_type = None
        if bt_raw:
            behavior_type = BEHAVIOR_TYPE_MAP.get(bt_raw.lower().strip())
            if not behavior_type:
                stats.errors += 1
                stats.error_details.append(
                    f"  Fila {row['_row']} ({nombre}): tipo_comportamiento '{bt_raw}' invalido. "
                    f"Valores: {', '.join(BEHAVIOR_TYPE_MAP.keys())}"
                )
                continue
        elif not padre_nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): tipo_comportamiento obligatorio para categorias raiz")
            continue

        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue

        payload = {"name": nombre}
        if behavior_type:
            payload["behavior_type"] = behavior_type
        if parent_id:
            payload["parent_id"] = parent_id

        status, resp = api.post("/api/v1/third-party-categories/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")

    if stats.skipped > 0 and not dry_run:
        st, resp = api.get("/api/v1/third-party-categories/", {"limit": 500})
        if st == 200:
            for item in resp.get("items", []):
                if item["name"] not in mapping:
                    mapping[item["name"]] = item["id"]

    return stats, mapping


def load_terceros(api, rows, tp_cat_map, dry_run):
    """Carga terceros asignando categorias por nombre (separadas por coma)."""
    stats = Stats()
    mapping = {}

    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue

        # Resolver categorias
        cats_raw = clean_str(row.get("categorias"))
        if not cats_raw:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): categorias es obligatorio")
            continue

        category_ids = []
        cat_errors = []
        for cat_name in [c.strip() for c in cats_raw.split(",") if c.strip()]:
            cat_id = find_in_map_ci(tp_cat_map, cat_name)
            if cat_id:
                category_ids.append(cat_id)
            else:
                cat_errors.append(cat_name)

        if cat_errors:
            stats.errors += 1
            stats.error_details.append(
                f"  Fila {row['_row']} ({nombre}): categorias no encontradas: {', '.join(cat_errors)}"
            )
            continue

        if dry_run:
            stats.created += 1
            mapping[nombre] = f"dry-run-{nombre}"
            continue

        payload = {
            "name": nombre,
            "initial_balance": parse_decimal(row.get("saldo_inicial"), 0),
            "category_ids": category_ids,
        }
        ident = clean_str(row.get("identificacion"))
        if ident:
            payload["identification_number"] = ident
        email = clean_str(row.get("email"))
        if email:
            payload["email"] = email
        tel = clean_str(row.get("telefono"))
        if tel:
            payload["phone"] = str(tel)
        addr = clean_str(row.get("direccion"))
        if addr:
            payload["address"] = addr

        status, resp = api.post("/api/v1/third-parties/", payload)
        if status in (200, 201):
            mapping[nombre] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): {status} - {resp}")

    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/third-parties/", rows, mapping)

    return stats, mapping


def load_materiales(api, rows, cat_map, un_map, dry_run):
    stats = Stats()
    mapping = {}
    for row in rows:
        codigo = clean_str(row.get("codigo"))
        nombre = clean_str(row.get("nombre"))
        categoria = clean_str(row.get("categoria"))
        unidad_negocio = clean_str(row.get("unidad_negocio"))

        errors = []
        if not codigo:
            errors.append("codigo es obligatorio")
        if not nombre:
            errors.append("nombre es obligatorio")

        cat_id = find_in_map_ci(cat_map, categoria) if categoria else None
        if not categoria:
            errors.append("categoria es obligatorio")
        elif not cat_id:
            errors.append(f"categoria '{categoria}' no encontrada")

        un_id = find_in_map_ci(un_map, unidad_negocio) if unidad_negocio else None
        if not unidad_negocio:
            errors.append("unidad_negocio es obligatorio")
        elif not un_id:
            errors.append(f"unidad_negocio '{unidad_negocio}' no encontrada")

        if errors:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo or '?'}): {'; '.join(errors)}")
            continue

        if dry_run:
            stats.created += 1
            mapping[codigo] = f"dry-run-{codigo}"
            continue

        payload = {
            "code": codigo,
            "name": nombre,
            "category_id": cat_id,
            "business_unit_id": un_id,
            "default_unit": clean_str(row.get("unidad")) or "kg",
        }
        desc = clean_str(row.get("descripcion"))
        if desc:
            payload["description"] = desc

        status, resp = api.post("/api/v1/materials/", payload)
        if status in (200, 201):
            mapping[codigo] = resp["id"]
            stats.created += 1
        elif _is_duplicate(status, resp):
            stats.skipped += 1
        else:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo}): {status} - {resp}")

    if stats.skipped > 0 and not dry_run:
        _fill_mapping_from_api(api, "/api/v1/materials/", rows, mapping, key_field="codigo", resp_key="code")

    return stats, mapping


def load_precios(api, rows, mat_map, dry_run):
    stats = Stats()
    for row in rows:
        codigo = clean_str(row.get("material_codigo"))
        if not codigo:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: material_codigo es obligatorio")
            continue

        mat_id = find_in_map_ci(mat_map, codigo)
        if not mat_id:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo}): material no encontrado")
            continue

        precio_compra = parse_decimal(row.get("precio_compra"), 0)
        precio_venta = parse_decimal(row.get("precio_venta"), 0)

        if precio_compra == 0 and precio_venta == 0:
            stats.skipped += 1
            continue

        if dry_run:
            stats.created += 1
            continue

        if precio_compra > 0:
            payload = {"material_id": mat_id, "purchase_price": precio_compra}
            notas = clean_str(row.get("notas"))
            if notas:
                payload["notes"] = notas
            status, resp = api.post("/api/v1/price-lists/", payload)
            if status in (200, 201):
                stats.created += 1
            else:
                stats.errors += 1
                stats.error_details.append(f"  Fila {row['_row']} ({codigo}) compra: {status} - {resp}")

        if precio_venta > 0:
            payload = {"material_id": mat_id, "sale_price": precio_venta}
            status, resp = api.post("/api/v1/price-lists/", payload)
            if status in (200, 201):
                stats.created += 1
            else:
                stats.errors += 1
                stats.error_details.append(f"  Fila {row['_row']} ({codigo}) venta: {status} - {resp}")

    return stats, {}


def load_inventario(api, rows, mat_map, bodega_map, dry_run, db_url=None, org_id=None):
    """Carga inventario inicial directo en BD (sin crear ajustes ni afectar P&L).

    Si un material aparece en multiples bodegas, suma las cantidades y usa
    promedio ponderado para el costo. Actualiza stock global del material.
    """
    stats = Stats()

    if not dry_run and not db_url:
        stats.errors += 1
        stats.error_details.append("  db_url requerido para carga de inventario (no dry-run)")
        return stats, {}

    # Agrupar por material: sumar cantidades, promedio ponderado de costo
    from collections import defaultdict
    grouped = defaultdict(lambda: {"qty": 0, "value": 0, "cost": 0, "rows": []})
    for row in rows:
        codigo = clean_str(row.get("material_codigo"))
        cantidad = parse_decimal(row.get("cantidad"), 0)
        costo = parse_decimal(row.get("costo_unitario"), 0)
        if not codigo or cantidad == 0:
            continue
        g = grouped[codigo]
        g["qty"] += cantidad
        g["value"] += abs(cantidad) * abs(costo)
        g["cost"] = costo  # fallback si solo hay 1 fila
        g["rows"].append(row)

    # Recalcular costo promedio ponderado por material
    consolidated_rows = []
    for codigo, g in grouped.items():
        if g["qty"] > 0 and g["value"] > 0:
            avg_cost = g["value"] / abs(g["qty"])
        elif g["qty"] < 0 and g["value"] > 0:
            avg_cost = g["value"] / abs(g["qty"])
        else:
            avg_cost = abs(g["cost"])
        consolidated_rows.append({
            "material_codigo": codigo,
            "cantidad": g["qty"],
            "costo_unitario": avg_cost,
            "bodega": clean_str(g["rows"][0].get("bodega")),
            "_row": g["rows"][0]["_row"],
        })
    rows = consolidated_rows

    # Conectar a BD directamente (no via API)
    db_conn = None
    if not dry_run:
        try:
            from sqlalchemy import create_engine, text
            engine = create_engine(db_url)
            db_conn = engine.connect()
        except Exception as e:
            stats.errors += 1
            stats.error_details.append(f"  Error conectando a BD: {e}")
            return stats, {}

    for row in rows:
        codigo = clean_str(row.get("material_codigo"))
        cantidad = parse_decimal(row.get("cantidad"), 0)
        costo = parse_decimal(row.get("costo_unitario"), 0)
        bodega_nombre = clean_str(row.get("bodega"))

        if not codigo:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: material_codigo es obligatorio")
            continue

        if cantidad == 0:
            stats.skipped += 1
            continue

        mat_id = find_in_map_ci(mat_map, codigo)
        if not mat_id:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo}): material no encontrado")
            continue

        bodega_id = find_in_map_ci(bodega_map, bodega_nombre) if bodega_nombre else None
        if not bodega_id:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo}): bodega '{bodega_nombre}' no encontrada")
            continue

        if dry_run:
            stats.created += 1
            continue

        try:
            # Actualizar material: stock y avg_cost (directo, sin crear ajuste ni movimiento)
            db_conn.execute(text("""
                UPDATE materials SET
                    current_stock = :qty,
                    current_stock_liquidated = :qty,
                    current_average_cost = :cost
                WHERE id = CAST(:mat_id AS uuid)
            """), {"qty": cantidad, "cost": abs(costo), "mat_id": mat_id})

            db_conn.commit()
            stats.created += 1
        except Exception as e:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({codigo}): DB error: {e}")

    if db_conn:
        db_conn.close()

    return stats, {}


def load_activos_fijos(api, rows, gastos_map, dry_run, db_url=None, org_id=None):
    """Carga activos fijos directo en BD (sin crear MoneyMovement ni afectar cuentas/terceros)."""
    stats = Stats()

    if not dry_run and not db_url:
        stats.errors += 1
        stats.error_details.append("  db_url requerido para carga de activos (no dry-run)")
        return stats, {}

    db_conn = None
    if not dry_run:
        try:
            from sqlalchemy import create_engine, text
            engine = create_engine(db_url)
            db_conn = engine.connect()
        except Exception as e:
            stats.errors += 1
            stats.error_details.append(f"  Error conectando a BD: {e}")
            return stats, {}

    for row in rows:
        nombre = clean_str(row.get("nombre"))
        if not nombre:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']}: nombre es obligatorio")
            continue

        valor_compra = parse_decimal(row.get("valor_compra"), 0)
        if valor_compra <= 0:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): valor_compra debe ser > 0")
            continue

        valor_residual = parse_decimal(row.get("valor_residual"), 0)
        tasa = parse_decimal(row.get("tasa_depreciacion"), 0)
        if tasa <= 0:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): tasa_depreciacion debe ser > 0")
            continue

        dep_acumulada = parse_decimal(row.get("depreciacion_acumulada"), 0)
        current_value = valor_compra - dep_acumulada

        fecha_compra = clean_str(row.get("fecha_compra"))
        fecha_inicio = clean_str(row.get("fecha_inicio_depreciacion"))
        if not fecha_compra or not fecha_inicio:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): fechas son obligatorias")
            continue

        cat_nombre = clean_str(row.get("categoria_gasto"))
        cat_id = find_in_map_ci(gastos_map, cat_nombre) if cat_nombre else None

        codigo = clean_str(row.get("codigo_activo"))
        notas = clean_str(row.get("notas"))

        # Calcular depreciacion mensual y vida util
        monthly_dep = round(valor_compra * (tasa / 100), 2)
        depreciable = valor_compra - valor_residual
        useful_life = int(depreciable / monthly_dep) if monthly_dep > 0 else 0

        # Determinar status
        if current_value <= valor_residual:
            status_val = "fully_depreciated"
        else:
            status_val = "active"

        if dry_run:
            stats.created += 1
            continue

        try:
            from sqlalchemy import text
            db_conn.execute(text("""
                INSERT INTO fixed_assets
                    (id, organization_id, name, asset_code, notes,
                     purchase_date, depreciation_start_date,
                     purchase_value, salvage_value, current_value, accumulated_depreciation,
                     depreciation_rate, monthly_depreciation, useful_life_months,
                     expense_category_id, status)
                VALUES
                    (gen_random_uuid(), CAST(:org_id AS uuid), :name, :code, :notes,
                     CAST(:purchase_date AS date), CAST(:dep_start AS date),
                     :purchase_value, :salvage_value, :current_value, :accumulated_dep,
                     :dep_rate, :monthly_dep, :useful_life,
                     CAST(:cat_id AS uuid), :status)
            """), {
                "org_id": org_id, "name": nombre, "code": codigo, "notes": notas,
                "purchase_date": fecha_compra, "dep_start": fecha_inicio,
                "purchase_value": valor_compra, "salvage_value": valor_residual,
                "current_value": current_value, "accumulated_dep": dep_acumulada,
                "dep_rate": tasa, "monthly_dep": monthly_dep, "useful_life": useful_life,
                "cat_id": cat_id, "status": status_val,
            })
            db_conn.commit()
            stats.created += 1
        except Exception as e:
            stats.errors += 1
            stats.error_details.append(f"  Fila {row['_row']} ({nombre}): DB error: {e}")

    if db_conn:
        db_conn.close()

    return stats, {}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Carga inicial de datos maestros desde Excel v2",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--file", required=True, help="Ruta al archivo Excel")
    parser.add_argument("--api-url", default="http://localhost:8000", help="URL base de la API")
    parser.add_argument("--email", required=True, help="Email del usuario administrador")
    parser.add_argument("--password", required=True, help="Contraseña del usuario")
    parser.add_argument("--org-id", required=True, help="UUID de la organización")
    parser.add_argument("--db-url", default=None,
                        help="URL de BD para carga directa de inventario. Ej: postgresql://admin:password@localhost:5432/reciclaje_db")
    parser.add_argument("--dry-run", action="store_true", help="Solo validar, no crear datos")
    args = parser.parse_args()

    print(f"\n{'=' * 60}")
    print(f"  CARGA INICIAL DE DATOS MAESTROS v2")
    print(f"  {'(MODO DRY-RUN - solo validación)' if args.dry_run else ''}")
    print(f"{'=' * 60}")
    print(f"\nArchivo: {args.file}")
    print(f"API:     {args.api_url}")
    print(f"Org:     {args.org_id}\n")

    try:
        wb = load_workbook(args.file, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: Archivo '{args.file}' no encontrado")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: No se pudo leer el archivo Excel: {e}")
        sys.exit(1)

    print(f"Hojas encontradas: {', '.join(wb.sheetnames)}\n")

    api = APIClient(args.api_url, args.org_id)
    if not args.dry_run:
        print("Autenticando...")
        if not api.login(args.email, args.password):
            print("ERROR: No se pudo autenticar.")
            sys.exit(1)
        print("  OK\n")
    else:
        print("(Dry-run: saltando autenticación)\n")

    all_stats = {}
    all_errors = []
    total_sheets = 11

    # 1. Unidades de Negocio
    print(f"[1/{total_sheets}] UnidadesNegocio...")
    rows = read_sheet(wb, "UnidadesNegocio")
    stats, un_map = load_unidades_negocio(api, rows, args.dry_run)
    all_stats["UnidadesNegocio"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("UnidadesNegocio"))

    # 2. Categorías de Materiales
    print(f"[2/{total_sheets}] CategoriaMateriales...")
    rows = read_sheet(wb, "CategoriaMateriales")
    stats, cat_map = load_categorias_material(api, rows, args.dry_run)
    all_stats["CategoriaMateriales"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("CategoriaMateriales"))

    # 3. Bodegas
    print(f"[3/{total_sheets}] Bodegas...")
    rows = read_sheet(wb, "Bodegas")
    stats, bodega_map = load_bodegas(api, rows, args.dry_run)
    all_stats["Bodegas"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Bodegas"))

    # 4. Cuentas
    print(f"[4/{total_sheets}] Cuentas...")
    rows = read_sheet(wb, "Cuentas")
    stats, _ = load_cuentas(api, rows, args.dry_run)
    all_stats["Cuentas"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Cuentas"))

    # 5. Categorías de Gastos
    print(f"[5/{total_sheets}] CategoriaGastos...")
    rows = read_sheet(wb, "CategoriaGastos")
    stats, gastos_map = load_categoria_gastos(api, rows, un_map, args.dry_run)
    all_stats["CategoriaGastos"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("CategoriaGastos"))

    # 6. Categorías de Terceros
    print(f"[6/{total_sheets}] CategoriaTerceros...")
    rows = read_sheet(wb, "CategoriaTerceros")
    stats, tp_cat_map = load_categoria_terceros(api, rows, args.dry_run)
    all_stats["CategoriaTerceros"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("CategoriaTerceros"))

    # 7. Terceros
    print(f"[7/{total_sheets}] Terceros...")
    rows = read_sheet(wb, "Terceros")
    stats, _ = load_terceros(api, rows, tp_cat_map, args.dry_run)
    all_stats["Terceros"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Terceros"))

    # 8. Materiales
    print(f"[8/{total_sheets}] Materiales...")
    rows = read_sheet(wb, "Materiales")
    stats, mat_map = load_materiales(api, rows, cat_map, un_map, args.dry_run)
    all_stats["Materiales"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Materiales"))

    # 9. Precios
    print(f"[9/{total_sheets}] Precios...")
    rows = read_sheet(wb, "Precios")
    stats, _ = load_precios(api, rows, mat_map, args.dry_run)
    all_stats["Precios"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Precios"))

    # 10. Inventario (stock inicial via adjustments/increase)
    print(f"[10/{total_sheets}] Inventario...")
    rows = read_sheet(wb, "Inventario")
    db_url = args.db_url
    if not db_url and not args.dry_run:
        # Intentar leer de .env o variable de entorno
        import os
        db_url = os.environ.get("DATABASE_URL")
        if not db_url:
            env_path = os.path.join(os.path.dirname(__file__), "..", ".env")
            if os.path.exists(env_path):
                with open(env_path) as f:
                    for line in f:
                        if line.startswith("DATABASE_URL="):
                            db_url = line.split("=", 1)[1].strip().strip('"').strip("'")
                            break
        if not db_url:
            print("  ⚠️  --db-url no proporcionado y DATABASE_URL no encontrada. Inventario no se cargara.")
    stats, _ = load_inventario(api, rows, mat_map, bodega_map, args.dry_run, db_url=db_url, org_id=args.org_id)
    all_stats["Inventario"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("Inventario"))

    # 11. Activos Fijos (directo en BD)
    print(f"[11/{total_sheets}] ActivosFijos...")
    rows = read_sheet(wb, "ActivosFijos")
    stats, _ = load_activos_fijos(api, rows, gastos_map, args.dry_run, db_url=db_url, org_id=args.org_id)
    all_stats["ActivosFijos"] = stats
    all_errors.extend(stats.error_details)
    print(stats.report("ActivosFijos"))

    wb.close()

    # Resumen
    total_created = sum(s.created for s in all_stats.values())
    total_skipped = sum(s.skipped for s in all_stats.values())
    total_errors = sum(s.errors for s in all_stats.values())

    print(f"\n{'=' * 60}")
    print(f"  RESUMEN")
    print(f"{'=' * 60}")
    print(f"  Creados:  {total_created}")
    print(f"  Omitidos: {total_skipped} (ya existian)")
    print(f"  Errores:  {total_errors}")

    if all_errors:
        print(f"\n  DETALLE DE ERRORES:")
        for err in all_errors:
            print(f"    {err}")

    print()

    if total_errors > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
