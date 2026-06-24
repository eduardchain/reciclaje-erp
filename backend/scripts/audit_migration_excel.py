#!/usr/bin/env python3
"""Auditoria de Excel de migracion ANTES de correr migrate_org.py.

Valida estructura, completitud y cruces entre hojas. Reporta issues en
plain Spanish para que el cliente entienda que corregir.

Disenado para atrapar todos los gotchas que descubrimos en la migracion
de Biogreen (2026-06-23):
- typos en nombres (Genral vs General)
- texto suelto en fila NOTAS que rompe el parsing
- referencias cruzadas a entidades no declaradas (MapeoUN → CategoriaMateriales)
- jerarquia >2 niveles en CategoriaGastos
- formato de fecha invalido en Inventario
- columna 'categorias' vacia en Terceros
- bodega = nombre material en Inventario (en vez de bodega real)

Uso:
    ./venv/bin/python scripts/audit_migration_excel.py <ruta-excel>

Exit codes:
    0 — sin errores bloqueantes (puede tener warnings)
    1 — hay errores bloqueantes, debe corregirse antes del dry-run
    2 — error al abrir el archivo
"""
from __future__ import annotations

import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: instala openpyxl primero: ./venv/bin/pip install openpyxl")
    sys.exit(2)


# Las 7 categorias de terceros que se crean automaticamente al crear org
DEFAULT_TP_CATEGORIES = {
    "proveedor material",
    "proveedor servicios",
    "cliente",
    "socios",
    "generico",
    "genérico",
    "provision",
    "provisión",
    "pasivo",
}

VALID_BEHAVIOR_TYPES = {
    "proveedor_material",
    "proveedor_servicios",
    "cliente",
    "inversionista",
    "generico",
    "provision",
    "pasivo",
}

EXPECTED_SHEETS = {
    "NOTAS",
    "UnidadesNegocio",
    "Bodegas",
    "CategoriaMateriales",
    "CategoriaGastos",
    "CategoriaTerceros",
    "MapeoUN",
    "Materiales",
    "Precios",
    "Terceros",
    "Cuentas",
    "Inventario",
    "ActivosFijos",
}


class Issue:
    BLOQUEANTE = "BLOQUEANTE"
    WARNING = "WARNING"
    OK = "OK"


class AuditReport:
    def __init__(self):
        self.issues: list[tuple[str, str, str]] = []  # (level, sheet, msg)
        self.stats: dict[str, dict] = defaultdict(dict)

    def err(self, sheet: str, msg: str):
        self.issues.append((Issue.BLOQUEANTE, sheet, msg))

    def warn(self, sheet: str, msg: str):
        self.issues.append((Issue.WARNING, sheet, msg))

    @property
    def errors(self) -> list[tuple[str, str, str]]:
        return [i for i in self.issues if i[0] == Issue.BLOQUEANTE]

    @property
    def warnings(self) -> list[tuple[str, str, str]]:
        return [i for i in self.issues if i[0] == Issue.WARNING]

    def print_summary(self):
        print()
        print("=" * 78)
        print(f"  REPORTE: {len(self.errors)} bloqueantes · {len(self.warnings)} warnings")
        print("=" * 78)
        if self.errors:
            print("\n🔴 ERRORES BLOQUEANTES (corregir antes del dry-run):")
            for _, sheet, msg in self.errors:
                print(f"  [{sheet}] {msg}")
        if self.warnings:
            print("\n🟡 WARNINGS (revisar pero no bloquean):")
            for _, sheet, msg in self.warnings:
                print(f"  [{sheet}] {msg}")
        if not self.errors and not self.warnings:
            print("\n✅ TODO OK — el Excel esta listo para correr el dry-run.\n")
            return
        if not self.errors:
            print("\n✅ Sin errores bloqueantes — puedes proceder al dry-run.")
            print("   Revisa los warnings con el cliente si tienen sentido para su negocio.\n")
            return
        print(f"\n❌ {len(self.errors)} errores bloqueantes. Corregir antes de continuar.\n")


def clean(v) -> str:
    """Convierte a string limpio (None → '')."""
    if v is None:
        return ""
    return str(v).strip()


def read_sheet(wb, name: str) -> tuple[list[str], list[dict]]:
    """Lee una hoja replicando la logica de migrate_org.py read_sheet().

    - Si fila 1 es 'NOTAS:' con SOLO la primera celda llena, header = fila 2.
    - Si fila 1 tiene cualquier texto en otra celda, header = fila 1 (lo cual es
      un BUG visible: la hoja queda mal-leida por el script).
    """
    if name not in wb.sheetnames:
        return [], []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return [], []
    first = rows[0]
    is_notes = bool(
        first and first[0]
        and isinstance(first[0], str)
        and first[0].strip().upper().startswith("NOTAS")
        and all(v is None for v in first[1:])
    )
    header_idx = 1 if is_notes else 0
    headers = [
        clean(h).lower().rstrip("*").strip() if h else f"col_{i}"
        for i, h in enumerate(rows[header_idx])
    ]
    data = []
    for r_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if all(v is None for v in row):
            continue
        d = {"_row": r_idx}
        for c, v in enumerate(row):
            if c < len(headers):
                d[headers[c]] = v
        data.append(d)
    return headers, data


def check_notas_rows(wb, report: AuditReport):
    """Verifica que filas NOTAS solo tengan texto en col 1. Si no, el parser falla."""
    for sheet_name in wb.sheetnames:
        if sheet_name == "NOTAS":
            continue
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        first_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if not first_row or not first_row[0] or not isinstance(first_row[0], str):
            continue
        if not first_row[0].strip().upper().startswith("NOTAS"):
            continue
        stray = [
            (c, v) for c, v in enumerate(first_row[1:], start=2)
            if v is not None and clean(v) != ""
        ]
        if stray:
            col_letters = ", ".join(chr(64 + c) for c, _ in stray)
            values = ", ".join(repr(v) for _, v in stray)
            report.err(
                sheet_name,
                f"Fila 1 (NOTAS) tiene texto en columnas {col_letters}: {values}. "
                f"Borralo — debe estar SOLO en columna A. Si no, el script lee mal "
                f"la hoja y queda 'vacia'.",
            )


def check_unidades_negocio(rows: list[dict], report: AuditReport) -> set[str]:
    names = set()
    if not rows:
        report.err("UnidadesNegocio", "Hoja vacia. Es obligatoria — declarar minimo 1 UN.")
        return names
    seen = set()
    for r in rows:
        name = clean(r.get("nombre"))
        if not name:
            report.err("UnidadesNegocio", f"Fila {r['_row']}: nombre vacio.")
            continue
        key = name.lower()
        if key in seen:
            report.err("UnidadesNegocio", f"Fila {r['_row']}: nombre duplicado '{name}'.")
            continue
        seen.add(key)
        names.add(name)
    return names


def check_bodegas(rows: list[dict], report: AuditReport) -> set[str]:
    names = set()
    if not rows:
        report.warn(
            "Bodegas",
            "Hoja vacia. El script creara una bodega 'Principal' por defecto.",
        )
        return {"Principal"}
    seen = set()
    for r in rows:
        name = clean(r.get("nombre"))
        if not name:
            report.err("Bodegas", f"Fila {r['_row']}: nombre vacio.")
            continue
        key = name.lower()
        if key in seen:
            report.err("Bodegas", f"Fila {r['_row']}: nombre duplicado '{name}'.")
            continue
        seen.add(key)
        names.add(name)
    return names


def check_categoria_materiales(rows: list[dict], report: AuditReport) -> set[str]:
    names = set()
    if not rows:
        report.err("CategoriaMateriales", "Hoja vacia. Es obligatoria.")
        return names
    seen = set()
    for r in rows:
        name = clean(r.get("nombre"))
        if not name:
            report.err("CategoriaMateriales", f"Fila {r['_row']}: nombre vacio.")
            continue
        key = name.lower()
        if key in seen:
            report.err("CategoriaMateriales", f"Fila {r['_row']}: nombre duplicado '{name}'.")
            continue
        seen.add(key)
        names.add(name)
    return names


def check_categoria_gastos(rows: list[dict], report: AuditReport) -> set[str]:
    """Valida nombres, padres existen, jerarquia max 2 niveles."""
    names = set()
    parents = {}  # name → parent (lowercased keys)
    for r in rows:
        name = clean(r.get("nombre"))
        padre = clean(r.get("padre"))
        if not name:
            report.err("CategoriaGastos", f"Fila {r['_row']}: nombre vacio.")
            continue
        names.add(name)
        if padre:
            parents[name.lower()] = padre.lower()

    name_lower_set = {n.lower() for n in names}
    for r in rows:
        name = clean(r.get("nombre"))
        padre = clean(r.get("padre"))
        if not name or not padre:
            continue
        if padre.lower() not in name_lower_set:
            report.err(
                "CategoriaGastos",
                f"Fila {r['_row']} ({name}): padre '{padre}' no existe como categoria. "
                f"Crearlo primero o quitar el valor de padre.",
            )
        elif padre.lower() in parents:
            abuelo = parents[padre.lower()]
            report.err(
                "CategoriaGastos",
                f"Fila {r['_row']} ({name}): jerarquia excede 2 niveles "
                f"({abuelo} > {padre} > {name}). Aplanar — el API solo permite "
                f"categoria > subcategoria.",
            )
    return names


def check_categoria_terceros(rows: list[dict], report: AuditReport) -> set[str]:
    """Valida extras (las 7 default ya vienen con la org)."""
    names = set()
    for r in rows:
        name = clean(r.get("nombre"))
        bt = clean(r.get("tipo_comportamiento"))
        padre = clean(r.get("padre"))
        if not name:
            report.err("CategoriaTerceros", f"Fila {r['_row']}: nombre vacio.")
            continue
        if name.lower() in DEFAULT_TP_CATEGORIES:
            report.warn(
                "CategoriaTerceros",
                f"Fila {r['_row']} ({name}): ya es categoria default — el script "
                f"la salta silenciosamente. Puedes borrar la fila.",
            )
            continue
        if not bt and not padre:
            report.err(
                "CategoriaTerceros",
                f"Fila {r['_row']} ({name}): tipo_comportamiento obligatorio "
                f"para categorias raiz.",
            )
            continue
        if bt and bt.lower() not in VALID_BEHAVIOR_TYPES:
            valid = ", ".join(sorted(VALID_BEHAVIOR_TYPES))
            report.err(
                "CategoriaTerceros",
                f"Fila {r['_row']} ({name}): tipo_comportamiento '{bt}' invalido. "
                f"Debe ser uno de: {valid}.",
            )
        names.add(name)
    return names


def check_mapeo_un(
    rows: list[dict],
    cm_names: set[str],
    un_names: set[str],
    report: AuditReport,
) -> dict[str, str]:
    """Verifica cruces de MapeoUN contra CategoriaMateriales y UnidadesNegocio."""
    mapping = {}
    cm_lower = {n.lower(): n for n in cm_names}
    un_lower = {n.lower(): n for n in un_names}
    for r in rows:
        cat = clean(r.get("categoria_material"))
        un = clean(r.get("unidad_negocio_default"))
        if not cat and not un:
            continue
        if not cat:
            report.warn("MapeoUN", f"Fila {r['_row']}: categoria_material vacia.")
            continue
        if cat.lower() not in cm_lower:
            report.err(
                "MapeoUN",
                f"Fila {r['_row']}: categoria_material '{cat}' no existe en "
                f"CategoriaMateriales. Disponibles: {sorted(cm_names) or '[vacio]'}.",
            )
            continue
        if not un:
            report.warn(
                "MapeoUN",
                f"Fila {r['_row']} ({cat}): unidad_negocio_default vacia.",
            )
            continue
        if un.lower() not in un_lower:
            report.err(
                "MapeoUN",
                f"Fila {r['_row']} ({cat}): unidad_negocio_default '{un}' no existe "
                f"en UnidadesNegocio. Disponibles: {sorted(un_names) or '[vacio]'}.",
            )
            continue
        mapping[cat] = un
    return mapping


def check_materiales(
    rows: list[dict],
    cm_names: set[str],
    un_names: set[str],
    mapeo: dict[str, str],
    report: AuditReport,
) -> set[str]:
    """Verifica codigo unico, categoria existe, UN resoluble."""
    codes = set()
    seen_codes = Counter()
    cm_lower = {n.lower() for n in cm_names}
    un_lower = {n.lower() for n in un_names}
    mapeo_lower = {k.lower(): v for k, v in mapeo.items()}

    for r in rows:
        code = clean(r.get("codigo"))
        name = clean(r.get("nombre"))
        cat = clean(r.get("categoria"))
        un_override = clean(r.get("unidad_negocio"))
        if not code:
            report.err("Materiales", f"Fila {r['_row']}: codigo vacio.")
            continue
        if not name:
            report.err("Materiales", f"Fila {r['_row']} ({code}): nombre vacio.")
        seen_codes[code.lower()] += 1
        codes.add(code)

        if not cat:
            report.err("Materiales", f"Fila {r['_row']} ({code}): categoria vacia.")
        elif cat.lower() not in cm_lower:
            report.err(
                "Materiales",
                f"Fila {r['_row']} ({code}): categoria '{cat}' no existe en "
                f"CategoriaMateriales.",
            )

        # Resolver UN: override en la fila, o default desde MapeoUN
        un_name = un_override or (mapeo_lower.get(cat.lower()) if cat else None)
        if not un_name:
            report.err(
                "Materiales",
                f"Fila {r['_row']} ({code}): no se puede resolver UN. "
                f"O llena 'unidad_negocio' en esta fila, o agrega un mapeo "
                f"en MapeoUN para categoria '{cat}'.",
            )
        elif un_name.lower() not in un_lower:
            report.err(
                "Materiales",
                f"Fila {r['_row']} ({code}): UN resuelta '{un_name}' no existe "
                f"en UnidadesNegocio.",
            )

    for code, cnt in seen_codes.items():
        if cnt > 1:
            report.err("Materiales", f"Codigo '{code}' duplicado x{cnt}.")
    return codes


def check_precios(
    rows: list[dict],
    material_codes: set[str],
    report: AuditReport,
):
    code_lower = {c.lower() for c in material_codes}
    skipped = []
    for r in rows:
        code = clean(r.get("material_codigo"))
        if not code:
            report.err("Precios", f"Fila {r['_row']}: material_codigo vacio.")
            continue
        if code.lower() not in code_lower:
            report.err(
                "Precios",
                f"Fila {r['_row']}: material '{code}' no existe en Materiales.",
            )
            continue
        pc = r.get("precio_compra")
        pv = r.get("precio_venta")
        pc_n = float(pc) if isinstance(pc, (int, float)) else 0
        pv_n = float(pv) if isinstance(pv, (int, float)) else 0
        if pc_n <= 0 and pv_n <= 0:
            skipped.append(code)
    if skipped:
        report.warn(
            "Precios",
            f"{len(skipped)} material(es) sin precio_compra ni precio_venta — "
            f"se omitiran silenciosamente: {', '.join(skipped[:10])}"
            f"{'...' if len(skipped) > 10 else ''}.",
        )


def check_terceros(
    rows: list[dict],
    extras: set[str],
    report: AuditReport,
):
    all_cats_lower = DEFAULT_TP_CATEGORIES | {e.lower() for e in extras}
    no_cat = 0
    seen_docs = Counter()
    total_balance = 0.0
    suspicious_names = []

    for r in rows:
        name = clean(r.get("nombre"))
        cats = clean(r.get("categorias"))
        doc = clean(r.get("identificacion"))
        bal = r.get("saldo_inicial")
        if not name:
            report.err("Terceros", f"Fila {r['_row']}: nombre vacio.")
            continue

        if not cats:
            report.err(
                "Terceros",
                f"Fila {r['_row']} '{name}': columna 'categorias' vacia (obligatorio). "
                f"Asignar al menos una categoria, ej: 'Proveedor Material' o 'Cliente'.",
            )
            no_cat += 1
        else:
            for c in [x.strip() for x in cats.split(",") if x.strip()]:
                if c.lower() not in all_cats_lower:
                    valid_list = sorted(DEFAULT_TP_CATEGORIES | extras)
                    report.err(
                        "Terceros",
                        f"Fila {r['_row']} '{name}': categoria '{c}' no existe. "
                        f"Validas: {valid_list}.",
                    )

        if doc:
            seen_docs[doc] += 1

        if isinstance(bal, (int, float)):
            total_balance += float(bal)

        # Heuristica: nombres en mayusculas que parecen MATERIALES, no terceros
        if name.upper() == name and len(name) > 2:
            bad_words = {"BATERIA", "PLASTICO", "PAPEL", "CARTON", "HIERRO", "VIDRIO", "RECICLAJE"}
            if name in bad_words or name.split()[0] in bad_words:
                suspicious_names.append(name)

    for doc, cnt in seen_docs.items():
        if cnt > 1:
            report.warn("Terceros", f"Documento '{doc}' aparece en {cnt} filas — posible duplicado.")

    if suspicious_names:
        report.warn(
            "Terceros",
            f"Nombres sospechosos que parecen materiales/conceptos en vez de personas/empresas: "
            f"{', '.join(suspicious_names[:5])}. Verificar con el cliente.",
        )


def check_cuentas(rows: list[dict], report: AuditReport):
    seen = set()
    for r in rows:
        name = clean(r.get("nombre"))
        if not name:
            report.err("Cuentas", f"Fila {r['_row']}: nombre vacio.")
            continue
        if name.lower() in seen:
            report.err("Cuentas", f"Fila {r['_row']}: nombre duplicado '{name}'.")
        seen.add(name.lower())


def check_inventario(
    rows: list[dict],
    material_codes: set[str],
    bodega_names: set[str],
    report: AuditReport,
):
    code_lower = {c.lower() for c in material_codes}
    bo_lower = {b.lower() for b in bodega_names}
    for r in rows:
        code = clean(r.get("material_codigo"))
        bo = clean(r.get("bodega"))
        fecha = r.get("fecha")
        if not code:
            report.err("Inventario", f"Fila {r['_row']}: material_codigo vacio.")
            continue
        if code.lower() not in code_lower:
            report.err(
                "Inventario",
                f"Fila {r['_row']}: material '{code}' no existe en Materiales.",
            )
        if bo and bo.lower() not in bo_lower:
            report.err(
                "Inventario",
                f"Fila {r['_row']} ({code}): bodega '{bo}' no existe. "
                f"Disponibles: {sorted(bodega_names)}.",
            )
        # Validar formato fecha
        if fecha is not None and fecha != "":
            if isinstance(fecha, str):
                try:
                    datetime.strptime(fecha, "%Y-%m-%d")
                except ValueError:
                    report.err(
                        "Inventario",
                        f"Fila {r['_row']} ({code}): fecha '{fecha}' no es AAAA-MM-DD. "
                        f"Ej valido: '2026-06-23'. O dejar vacio para usar fecha de hoy.",
                    )


def check_activos_fijos(
    rows: list[dict],
    cg_names: set[str],
    un_names: set[str],
    report: AuditReport,
):
    cg_lower = {n.lower() for n in cg_names}
    un_lower = {n.lower() for n in un_names}
    for r in rows:
        name = clean(r.get("nombre"))
        valor = r.get("valor_compra")
        vida = r.get("vida_util_meses")
        cat = clean(r.get("categoria_gasto"))
        un = clean(r.get("unidad_negocio"))
        if not name:
            report.err("ActivosFijos", f"Fila {r['_row']}: nombre vacio.")
            continue
        # Heuristica: sample row del template que el cliente olvido borrar
        if "demo" in name.lower() or "ejemplo" in name.lower():
            report.warn(
                "ActivosFijos",
                f"Fila {r['_row']} ({name}): nombre parece ejemplo del template "
                f"('demo'/'ejemplo'). ¿El cliente tiene este activo o se olvido de borrarlo?",
            )
        if valor is None or not isinstance(valor, (int, float)):
            report.err("ActivosFijos", f"Fila {r['_row']} ({name}): valor_compra obligatorio.")
        if vida is None or not isinstance(vida, (int, float)):
            report.err(
                "ActivosFijos",
                f"Fila {r['_row']} ({name}): vida_util_meses obligatorio.",
            )
        if cat and cat.lower() not in cg_lower:
            report.err(
                "ActivosFijos",
                f"Fila {r['_row']} ({name}): categoria_gasto '{cat}' no existe en "
                f"CategoriaGastos.",
            )
        if un and un.lower() not in un_lower:
            report.err(
                "ActivosFijos",
                f"Fila {r['_row']} ({name}): unidad_negocio '{un}' no existe en "
                f"UnidadesNegocio.",
            )


def audit(excel_path: Path) -> AuditReport:
    report = AuditReport()
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"ERROR abriendo Excel: {e}")
        sys.exit(2)

    # Verificar hojas esperadas
    actual = set(wb.sheetnames)
    missing = EXPECTED_SHEETS - actual
    if missing:
        for s in missing:
            report.err("ESTRUCTURA", f"Falta la hoja '{s}'.")

    # Verificar fila NOTAS (texto suelto rompe el parser)
    check_notas_rows(wb, report)

    # Leer todas las hojas con la misma logica del script principal
    _, un_rows = read_sheet(wb, "UnidadesNegocio")
    _, bo_rows = read_sheet(wb, "Bodegas")
    _, cm_rows = read_sheet(wb, "CategoriaMateriales")
    _, cg_rows = read_sheet(wb, "CategoriaGastos")
    _, ct_rows = read_sheet(wb, "CategoriaTerceros")
    _, mu_rows = read_sheet(wb, "MapeoUN")
    _, ma_rows = read_sheet(wb, "Materiales")
    _, pr_rows = read_sheet(wb, "Precios")
    _, tp_rows = read_sheet(wb, "Terceros")
    _, cu_rows = read_sheet(wb, "Cuentas")
    _, in_rows = read_sheet(wb, "Inventario")
    _, af_rows = read_sheet(wb, "ActivosFijos")

    # Resumen
    print(f"Hojas: {sorted(actual)}")
    print(f"\nResumen de filas:")
    for s, d in [
        ("UnidadesNegocio", un_rows), ("Bodegas", bo_rows),
        ("CategoriaMateriales", cm_rows), ("CategoriaGastos", cg_rows),
        ("CategoriaTerceros", ct_rows), ("MapeoUN", mu_rows),
        ("Materiales", ma_rows), ("Precios", pr_rows),
        ("Terceros", tp_rows), ("Cuentas", cu_rows),
        ("Inventario", in_rows), ("ActivosFijos", af_rows),
    ]:
        print(f"  {s:25s} {len(d):4d} filas")

    # Validaciones encadenadas
    un_names = check_unidades_negocio(un_rows, report)
    bo_names = check_bodegas(bo_rows, report)
    cm_names = check_categoria_materiales(cm_rows, report)
    cg_names = check_categoria_gastos(cg_rows, report)
    ct_extras = check_categoria_terceros(ct_rows, report)
    mapeo = check_mapeo_un(mu_rows, cm_names, un_names, report)
    ma_codes = check_materiales(ma_rows, cm_names, un_names, mapeo, report)
    check_precios(pr_rows, ma_codes, report)
    check_terceros(tp_rows, ct_extras, report)
    check_cuentas(cu_rows, report)
    check_inventario(in_rows, ma_codes, bo_names, report)
    check_activos_fijos(af_rows, cg_names, un_names, report)

    return report


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"ERROR: archivo no existe: {excel_path}")
        sys.exit(2)

    print("=" * 78)
    print(f"  AUDITORIA — {excel_path.name}")
    print("=" * 78)
    report = audit(excel_path)
    report.print_summary()
    sys.exit(1 if report.errors else 0)


if __name__ == "__main__":
    main()
