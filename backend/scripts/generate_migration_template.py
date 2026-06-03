"""
Genera el template Excel unificado para migracion de organizaciones a EcoBalance.

Estructura: 1 hoja NOTAS + 12 hojas de datos con cabeceras + filas de ejemplo.
Las filas de ejemplo van con fondo AMARILLO para que el cliente las borre antes
de llenar con sus datos reales.

Uso (desde backend/):
    ./venv/bin/python scripts/generate_migration_template.py

Output: ../data/migration_template.xlsx
"""
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("ERROR: openpyxl no esta instalado.")
    sys.exit(1)


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NOTES_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SECTION_FONT = Font(bold=True, size=11, color="1F4E78")


SHEETS = [
    # (sheet_name, headers, example_rows, notes)
    (
        "UnidadesNegocio",
        ["nombre", "descripcion", "is_active"],
        [
            ["Chatarra", "Operacion de chatarra ferrosa", True],
            ["Metales", "Metales no ferrosos (aluminio, cobre, bronce)", True],
            ["Plastico", "Plasticos PET, PEAD, etc.", True],
        ],
        "Unidades de negocio (UN) — se cargan PRIMERO. Referenciadas por Materiales, "
        "CategoriaGastos y ActivosFijos. nombre es obligatorio. is_active default true.",
    ),
    (
        "Bodegas",
        ["nombre", "descripcion", "address", "is_active"],
        [
            ["Principal", "Bodega principal de operacion", "", True],
            ["Sucursal Norte", "Bodega secundaria zona norte", "", True],
        ],
        "Bodegas / almacenes. Si la hoja esta vacia se crea una bodega default 'Principal'. "
        "Las bodegas se referencian por nombre en la hoja Inventario.",
    ),
    (
        "CategoriaMateriales",
        ["nombre", "descripcion"],
        [
            ["FERROSOS", "Hierro / chatarra ferrosa"],
            ["NO FERROSOS", "Aluminio, cobre, bronce, etc."],
            ["PLASTICOS", "PET, PEAD, PVC, etc."],
        ],
        "Categorias de materiales. Una fila por categoria. nombre es obligatorio. "
        "Referenciadas por Materiales y por MapeoUN (UN default por categoria).",
    ),
    (
        "CategoriaGastos",
        [
            "nombre",
            "padre",
            "es_directo",
            "asignacion_un",
            "unidad_negocio",
            "unidades_negocio",
            "descripcion",
        ],
        [
            ["Logistica", "", "SI", "compartido", "", "Chatarra,Metales", "Gastos logisticos compartidos"],
            ["Combustible", "Logistica", "", "", "", "", "Subcategoria de Logistica"],
            ["Administracion", "", "NO", "general", "", "", "Gastos administrativos generales"],
            ["Servicios Publicos", "Administracion", "", "", "", "", "Subcategoria de Administracion"],
        ],
        "Categorias de gastos. Carga en 2-pass: primero raices (padre vacio), luego hijas. "
        "Max 2 niveles. Subcategorias heredan es_directo del padre (dejar en blanco). "
        "asignacion_un: directo|compartido|general (default general). "
        "unidad_negocio: UN unica. unidades_negocio: CSV cuando asignacion=compartido.",
    ),
    (
        "CategoriaTerceros",
        ["nombre", "tipo_comportamiento", "padre", "descripcion"],
        [
            ["Obligaciones Financieras", "inversionista", "", "Creditos bancarios, leasings, obligaciones de largo plazo"],
        ],
        "OPCIONAL: las 7 categorias default ya vienen con la org "
        "(Proveedor Material, Proveedor Servicios, Cliente, Socios, Generico, Provision, Pasivo). "
        "Solo agregar EXTRAS aqui. tipo_comportamiento: proveedor_material | proveedor_servicios | "
        "cliente | inversionista | generico | provision | pasivo. Solo en raices.",
    ),
    (
        "MapeoUN",
        ["categoria_material", "unidad_negocio_default", "notas"],
        [
            ["FERROSOS", "Chatarra", "Todo lo ferroso va a UN Chatarra"],
            ["NO FERROSOS", "Metales", "Aluminio, cobre, bronce → UN Metales"],
            ["PLASTICOS", "Plastico", "Todos los plasticos → UN Plastico"],
        ],
        "Mapeo de referencia (no se escribe al API). Sirve de FALLBACK cuando un material "
        "en la hoja Materiales no especifica su UN. Los materiales que necesiten otra UN "
        "deben sobreescribirla explicitamente en la columna unidad_negocio de Materiales.",
    ),
    (
        "Materiales",
        ["codigo", "nombre", "categoria", "unidad_negocio", "unidad", "descripcion"],
        [
            ["CH-001", "Chatarra Liviana", "FERROSOS", "Chatarra", "kg", ""],
            ["AL-001", "Aluminio Perfil", "NO FERROSOS", "Metales", "kg", ""],
            ["PET-001", "PET Botella", "PLASTICOS", "Plastico", "kg", ""],
        ],
        "Materiales del catalogo. codigo y nombre obligatorios. unidad_negocio vacio = usa "
        "default de MapeoUN. unidad: kg | ton | unit (default kg). "
        "current_stock y current_average_cost inician en 0 — se siembran via hoja Inventario.",
    ),
    (
        "Precios",
        ["material_codigo", "precio_compra", "precio_venta", "notas"],
        [
            ["CH-001", 1500, 2000, "Precios al corte"],
            ["AL-001", 6500, 8500, ""],
            ["PET-001", 800, 1200, ""],
        ],
        "Lista de precios inicial. Para cada fila se hacen 2 POST a /price-lists "
        "(uno purchase, otro sale). notas se envia a ambos. Filas con ambos precios "
        "en 0/vacio se omiten.",
    ),
    (
        "Terceros",
        [
            "nombre",
            "identificacion",
            "email",
            "telefono",
            "direccion",
            "categorias",
            "saldo_inicial",
            "provision_type",
        ],
        [
            ["Proveedor Demo SA", "900123456", "info@demo.com", "3001234567", "Calle 1 #2-3", "Proveedor Material", 0, ""],
            ["Cliente Demo Ltda", "901234567", "", "3009876543", "", "Cliente", 0, ""],
            ["Socio Demo", "", "", "", "", "Socios", 0, ""],
        ],
        "Terceros. categorias: CSV de nombres (M:N, ej: 'Proveedor Material,Cliente'). "
        "saldo_inicial: + = nos debe, - = le debemos. Sienta current_balance al crear "
        "(NO genera MoneyMovements de apertura). provision_type solo si tiene categoria Provision.",
    ),
    (
        "Cuentas",
        ["nombre", "tipo", "saldo_inicial", "numero_cuenta", "banco"],
        [
            ["Caja Principal", "efectivo", 0, "", ""],
            ["Banco Demo Cuenta Corriente", "banco", 0, "", ""],
        ],
        "Cuentas de dinero. tipo: efectivo|banco|digital (se normaliza a cash/bank/digital). "
        "saldo_inicial >= 0 obligatorio — los sobregiros se modelan como Tercero categoria 'Pasivo'.",
    ),
    (
        "Inventario",
        ["material_codigo", "bodega", "cantidad", "costo_unitario", "fecha"],
        [
            ["CH-001", "Principal", 500, 1200, ""],
        ],
        "Stock inicial por material+bodega. Solo materiales con stock > 0. "
        "Se siembra via POST /inventory/adjustments/increase con reason='Carga inicial migracion {org_name}' "
        "(marcador de decision #28 — excluye estos ajustes del P&L). "
        "fecha llena MaterialCostHistory.transaction_date para Balance Detallado historico exacto.",
    ),
    (
        "ActivosFijos",
        [
            "nombre",
            "codigo_activo",
            "fecha_compra",
            "fecha_inicio_depreciacion",
            "valor_compra",
            "valor_residual",
            "vida_util_meses",
            "depreciacion_acumulada",
            "categoria_gasto",
            "proveedor",
            "cuenta_pago",
            "unidad_negocio",
            "unidades_negocio",
            "notas",
        ],
        [
            ["Bascula Demo", "ACT-001", "2024-01-01", "2024-01-01", 5000000, 500000, 120, 750000, "Logistica", "", "", "", "", "Ejemplo carga historica (proveedor y cuenta_pago vacios → historical_load)"],
        ],
        "Activos fijos. depreciation_rate se calcula como 100/vida_util_meses. "
        "XOR de pago: SI proveedor → asset_purchase (mueve balance proveedor). "
        "SI cuenta_pago → asset_payment (mueve cuenta). "
        "SI AMBOS VACIOS → historical_load=true (no MoneyMovement, no cambia balances — para activos depreciados ya pagados en sistema origen). "
        "Validacion: depreciacion_acumulada <= valor_compra - valor_residual. "
        "categoria_gasto debe existir en hoja CategoriaGastos. "
        "unidad_negocio (single) o unidades_negocio (CSV) opcionales.",
    ),
]


NOTES_SHEET_CONTENT = [
    ("title", "Template de migracion EcoBalance"),
    ("blank", ""),
    ("section", "Como usar este template"),
    ("text", "1. Las filas en AMARILLO son ejemplos. BORRAR TODAS las filas amarillas antes de llenar con datos reales."),
    ("text", "2. La fila 2 de cada hoja es la cabecera (azul). No modificarla."),
    ("text", "3. Las hojas se cargan en orden — respetar dependencias (UN antes de Materiales, etc.)."),
    ("text", "4. Validar con --dry-run antes de --apply."),
    ("blank", ""),
    ("section", "Hojas (orden de carga)"),
    ("text", "1. UnidadesNegocio — UN del negocio (Chatarra, Metales, etc.). Se cargan PRIMERO."),
    ("text", "2. Bodegas — almacenes. Si esta vacia se crea 'Principal' por default."),
    ("text", "3. CategoriaMateriales — categorias del catalogo (FERROSOS, etc.)."),
    ("text", "4. CategoriaGastos — categorias de gastos, max 2 niveles, con asignacion UN."),
    ("text", "5. CategoriaTerceros — SOLO categorias EXTRA. Las 7 default (Proveedor Material, Cliente, Socios, etc.) ya vienen con la org."),
    ("text", "6. MapeoUN — referencia: UN default por categoria de material (fallback)."),
    ("text", "7. Materiales — catalogo de materiales con codigo + UN."),
    ("text", "8. Precios — lista de precios inicial (compra y venta) por material."),
    ("text", "9. Terceros — proveedores, clientes, socios, etc. con saldo inicial."),
    ("text", "10. Cuentas — cuentas de dinero (caja, banco, digital). Saldo >= 0."),
    ("text", "11. Inventario — stock inicial por material+bodega."),
    ("text", "12. ActivosFijos — activos fijos con depreciacion acumulada historica."),
    ("blank", ""),
    ("section", "Notas importantes"),
    ("text", "* ActivosFijos: si proveedor Y cuenta_pago vacios → historical_load=true (no genera MoneyMovement, no mueve balances). Para activos ya pagados en sistema origen."),
    ("text", "* Inventario: se carga via /inventory/adjustments/increase con reason='Carga inicial migracion {org}' — marcador de decision #28 que excluye estos ajustes del P&L (es patrimonio absorbido por los socios, no ganancia operativa)."),
    ("text", "* Terceros: saldo_inicial setea current_balance al crear el tercero (NO genera MoneyMovements de apertura). + = nos debe, - = le debemos."),
    ("text", "* Cuentas: saldo_inicial >= 0 obligatorio. Sobregiros se modelan como Tercero categoria 'Pasivo'."),
    ("text", "* CategoriaTerceros: NO duplicar las 7 default. Solo agregar extras (ej: 'Obligaciones Financieras' con tipo=inversionista)."),
    ("text", "* Org + admin + 5 roles + 7 categorias default + 71 permisos se siembran AUTOMATICAMENTE al crear la org via /system/organizations."),
    ("blank", ""),
    ("section", "Verificacion post-carga"),
    ("text", "* count(materials) >= rows en hoja Materiales"),
    ("text", "* sum(money_accounts.balance) == sum(Excel saldo_inicial) ± balance-tolerance"),
    ("text", "* sum(fixed_assets.current_value) == sum(valor_compra - depreciacion_acumulada) ± tolerance"),
    ("text", "* Balance Sheet is_balanced=true (activos == pasivos + patrimonio)"),
]


def build_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="NOTAS", index=0)
    ws.column_dimensions["A"].width = 140
    row_idx = 1
    for kind, text in NOTES_SHEET_CONTENT:
        cell = ws.cell(row=row_idx, column=1, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
            cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row_idx].height = 28
        elif kind == "section":
            cell.font = SECTION_FONT
            cell.fill = NOTES_FILL
            ws.row_dimensions[row_idx].height = 22
        elif kind == "text":
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # blank → just bump row
        row_idx += 1


def build_workbook() -> Workbook:
    wb = Workbook()
    # Eliminar sheet default
    default_sheet = wb.active
    wb.remove(default_sheet)

    # NOTAS al inicio
    build_notes_sheet(wb)

    for sheet_name, headers, example_rows, notes in SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        # Fila 1: notas
        ws.cell(row=1, column=1, value=f"NOTAS: {notes}").fill = NOTES_FILL
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        notes_cell = ws.cell(row=1, column=1)
        notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 55
        # Fila 2: cabecera
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        # Filas 3+: ejemplos (AMARILLO para indicar BORRAR)
        for row_idx, example in enumerate(example_rows, start=3):
            for col_idx, value in enumerate(example, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = EXAMPLE_FILL
        # Ajustar ancho de columnas
        for col_idx, header in enumerate(headers, start=1):
            col_letter = ws.cell(row=2, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = max(18, len(header) + 4)
        ws.freeze_panes = "A3"
    return wb


def main():
    out_path = Path(__file__).resolve().parent.parent.parent / "data" / "migration_template.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(out_path)
    print(f"Template generado: {out_path}")
    print(f"  Hojas: NOTAS, {', '.join(name for name, *_ in SHEETS)}")


if __name__ == "__main__":
    main()
